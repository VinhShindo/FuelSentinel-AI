#!/usr/bin/env python3
import os
import sys
import io
import json
import sqlite3
import threading
import logging
import traceback
import urllib.parse
from functools import wraps
from datetime import datetime, timedelta
from collections import deque
import random
import smtplib
import ssl
from email.message import EmailMessage

import certifi
import numpy as np
import pandas as pd
import torch
from scipy import signal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from flask import Flask, render_template, jsonify, send_file, request, redirect, url_for, flash, session
from werkzeug.security import generate_password_hash, check_password_hash
from supabase import create_client, Client
from dotenv import load_dotenv
from fastapi import FastAPI, Request as FastAPIRequest
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response

load_dotenv()

GMAIL_EMAIL = os.getenv("GMAIL_EMAIL")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
FLASK_SECRET_KEY = os.getenv("FLASK_SECRET_KEY", "fuel_sentinel_secret_key")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise RuntimeError("Thiếu SUPABASE_URL hoặc SUPABASE_KEY trong file .env")

log = logging.getLogger('werkzeug')
log.setLevel(logging.WARNING)
logger = logging.getLogger('FuelSentinel')
logger.setLevel(logging.INFO)
logger.propagate = False
logger.handlers.clear()

console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

SRC_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_DIR = os.path.dirname(SRC_DIR)
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
if SRC_DIR not in sys.path:
    sys.path.insert(0, SRC_DIR)
EVENT_LOG_DIR = os.path.join(BASE_DIR, 'data', 'logs')
os.makedirs(EVENT_LOG_DIR, exist_ok=True)

runtime_log_path = os.path.join(EVENT_LOG_DIR, 'runtime_terminal.log')
runtime_file_handler = logging.FileHandler(runtime_log_path, encoding='utf-8')
runtime_file_handler.setLevel(logging.INFO)
runtime_file_handler.setFormatter(formatter)
logger.addHandler(runtime_file_handler)

sys.path.insert(0, SRC_DIR)

WINDOW_SIZE = 10
BUFFER_SIZE = 120
MAX_POINTS_HISTORY = 100
DEFAULT_VISIBLE_POINTS = 20
CANDIDATE_THRESH = 0.55
CONFIRM_THRESH = 0.75
CONFIRM_COUNT = 2
FINISH_THRESH = 0.5
CANDIDATE_MISMATCH_TOLERANCE = 1
FUEL_EVENT_MISMATCH_TOLERANCE = 0
REFUEL_MIN_DELTA = 3.0
THEFT_MIN_DELTA = 1.5
DRIVING_MIN_FUEL_DELTA = 3.0
IDLE_MIN_FUEL_DELTA = 2.0
DRIVING_SPEED_GATE = 8.0
WEAK_EVENT_DURATION_GRACE_S = 900
IDLE_FUEL_SLOPE_RESET = 0.35
BACKTRACK_MAX_LOOKBACK = 20
BACKTRACK_FUEL_THRESHOLD = 0.2

EVENT_LOG_PATH = os.path.join(EVENT_LOG_DIR, "event_log.jsonl")
TOAST_LOG_PATH = os.path.join(EVENT_LOG_DIR, "toast_log.jsonl")
REALTIME_DB_PATH = os.path.join(EVENT_LOG_DIR, "realtime.db")
os.makedirs(EVENT_LOG_DIR, exist_ok=True)

MODEL_PATH = os.path.join(BASE_DIR, "outputs", "cnn_gru_20260815_075909", "final_model.pt")
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")
LABELS = ['Driving', 'Idle', 'Refuel', 'Theft']
FUSION_CSV = os.path.join(BASE_DIR, "data", "processed", "fusion", "fusion_dataset.csv")

FEATURE_COLS_21 = [
    "duration_min", "sample_count", "fuel_start", "fuel_end",
    "fuel_change", "fuel_mean", "fuel_std", "fuel_slope",
    "trend_r2", "trend_rmse", "max_drop", "max_rise",
    "drop_count", "rise_count", "fuel_range", "fuel_mad",
    "oscillation_count", "total_distance", "speed_mean",
    "speed_max", "speed_zero_ratio"
]

WINDOW_RANGES = {
    'normal':    (6, 8),
    'candidate': (4, 5),
    'Refuel':    (3, 4),
    'Theft':     (2, 4),
    'Idle':      (4, 7),
    'Driving':   (5, 8),
}
IDLE_STEPS = [4, 5, 6, 8]
STABILITY_STD_THRESHOLD = 0.10

from trainning.models.all_models import CNNGRUClassifier
from trainning.datasets.dataset_builder import convert_gps_to_relative
from feature_extraction.feature_extraction import build_segment_features_table

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

API_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__,
            template_folder=os.path.join(API_DIR, 'templates'),
            static_folder=os.path.join(API_DIR, 'static'))
app.secret_key = FLASK_SECRET_KEY

fastapi_app = FastAPI(
    title="FuelSentinel FastAPI",
    version="1.0.0",
    description="Backend API for FuelSentinel while keeping the Flask UI as the front-end entrypoint.",
)
fastapi_app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:5000",
        "http://localhost:5000",
        "http://127.0.0.1:8080",
        "http://localhost:8080",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@fastapi_app.get("/health")
def fastapi_health():
    return {
        "status": "ok",
        "service": "FuelSentinel FastAPI",
        "timestamp": datetime.utcnow().isoformat() + "Z",
    }


@fastapi_app.api_route("/api/{path:path}", methods=["GET", "POST", "PUT", "PATCH", "DELETE"])
async def fastapi_backend_proxy(request: FastAPIRequest, path: str):
    client = app.test_client()
    full_path = f"/api/{path.lstrip('/')}"
    query_params = dict(request.query_params)
    if query_params:
        full_path = f"{full_path}?{urllib.parse.urlencode(query_params)}"

    body = await request.body()
    headers = {key: value for key, value in request.headers.items() if key.lower() not in {"host", "content-length"}}
    response = client.open(
        full_path,
        method=request.method,
        data=body if body else None,
        headers=headers,
        environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
    )

    if response.is_json:
        return JSONResponse(status_code=response.status_code, content=response.get_json())
    return Response(status_code=response.status_code, content=response.get_data(), media_type=response.mimetype or "application/json")


def run_fastapi_backend(host: str = "127.0.0.1", port: int = 8000):
    import uvicorn

    uvicorn.run(app=fastapi_app, host=host, port=port, reload=False)


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized"}), 401
            flash("Vui lòng đăng nhập trước!", "danger")
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped_view


class KalmanFilter:
    def __init__(self, process_noise=0.02, measurement_noise=0.1, initial_value=0.0):
        self.process_noise = process_noise
        self.measurement_noise = measurement_noise
        self.x = initial_value
        self.p = 1.0
        self.k = 0.0
        self.initialized = False

    def update(self, measurement):
        if not self.initialized:
            self.x = measurement
            self.p = 1.0
            self.initialized = True
            return measurement
        self.p = self.p + self.process_noise
        self.k = self.p / (self.p + self.measurement_noise)
        self.x = self.x + self.k * (measurement - self.x)
        self.p = (1 - self.k) * self.p
        return self.x

    def reset(self):
        self.x = 0.0
        self.p = 1.0
        self.k = 0.0
        self.initialized = False

    def anchor(self, value, p=5.0):
        self.x = value
        self.p = p
        self.initialized = True


class AdaptiveNoiseFilter:
    """
    Tang loc nhieu du lieu fuel/speed - TACH BIET hoan toan voi tang nhan
    dien su kien (Refuel/Theft/Idle/Driving). Tu phat hien thay doi that
    trong chinh du lieu (khong can doi tracker/classifier "bat den xanh").

    Kien truc:
      1) Persistence check tren RAW: 2 delta lien tiep cung chieu, du lon
         (>= EVENT_DELTA_MIN) -> coi la THAY DOI THAT (refuel/theft/level-
         shift), tin tuong raw hoan toan va cho Kalman bam nhanh.
      2) Neu KHONG persistent nhung la spike (Hampel median/MAD) -> thay
         the bang median cua lich su GAN NHAT - nhung CHI toi da
         MAX_FREEZE tick lien tiep. Day la diem sua loi quan trong nhat so
         voi ban cu: ban cu dung `value_for_kalman = prev_filtered` khi
         khong persistent, khien filter DONG BANG VO THOI HAN (backtest
         cho thay sai lech toi +56L, keo dai hang chuc diem lien tuc).
         Ban moi khong bao gio dong bang qua MAX_FREEZE tick: sau do bat
         buoc chap nhan raw la thay doi that.
      3) Kalman filter lam muot phan nhieu Gaussian/quantization con lai,
         voi process noise TANG khi persistence duoc xac nhan (bam nhanh
         theo su kien that) va THAP trong trang thai binh thuong (nen
         nhieu jitter).
    """
    RAW_HISTORY_LEN = 15
    EVENT_DELTA_MIN = 3.0
    REFUEL_START_JUMP = 40.0
    THEFT_JUMP = 8.0
    SPIKE_JUMP = 18.0
    DRIVING_SPIKE_JUMP = 8.0
    MAX_REFUEL_STEP = 75.0
    STATIONARY_SPIKE_JUMP = 8.0
    MAX_JUMP = 45.0
    EMA_ALPHA = 0.90
    SPEED_BUFFER_LEN = 5
    RAMP_MIN_STEPS = 2
    RAMP_DEVIATION_MIN = 12.0
    RAMP_DEVIATION_RATIO = 0.20
    POST_REFUEL_HANDOFF_TOLERANCE = 6.0
    POST_REFUEL_HOLD_TICKS = 3

    def __init__(self, kalman_pn=0.08, kalman_mn=0.3):
        # kalman_pn/kalman_mn giu lai de tuong thich chu ky khoi tao cu (khong bat buoc dung)
        self.kalman = KalmanFilter(process_noise=kalman_pn, measurement_noise=kalman_mn)
        self.raw_buffer = deque(maxlen=self.RAW_HISTORY_LEN)         # {'fuel','speed'} - dung o app.py khac
        self.filtered_history = deque(maxlen=self.RAW_HISTORY_LEN)
        self.speed_buffer = deque(maxlen=self.SPEED_BUFFER_LEN)
        self.tracker_status = 'normal'   # giu de tuong thich set_tracker_status(), KHONG con
        self.event_type = None           # dung de quyet dinh q/r ben trong filter_fuel nua
        self.TANK_CAPACITY = 450.0
        self.current_window = 8
        self._last_flag = None
        self._pending_jump = None
        self._retroactive_point = None
        self._refuel_ramp = False
        self._theft_ramp = False
        self._refuel_increments = deque(maxlen=4)
        self._refuel_level = None
        self._post_refuel_hold_ticks = 0

    def set_tracker_status(self, status, event_type=None):
        # Giu nguyen chu ky goi cu de tuong thich voi phan con lai cua app.py,
        # nhung chi de hien thi/debug - tang loc nhieu tu phat hien thay doi
        # that qua persistence-check ngay ben duoi, doc lap voi ket qua nhan nhan.
        self.tracker_status = status
        self.event_type = event_type

    def _spike_gate(self):
        return self.SPIKE_JUMP

    def _baseline(self):
        if not self.filtered_history:
            return None
        return float(self.filtered_history[-1])

    def _is_stationary_event(self, delta, speed):
        # A single downward jump while stationary is ambiguous: it can be a
        # theft, but it can also be an isolated sensor outlier. Refuel has a
        # much stronger physical signature and may still start immediately.
        return speed <= 5.0 and delta >= self.REFUEL_START_JUMP

    def _is_pending_spike(self, delta, speed):
        if speed > 5.0:
            return abs(delta) >= self.DRIVING_SPIKE_JUMP
        return abs(delta) >= self.STATIONARY_SPIKE_JUMP

    def filter_fuel(self, fuel_raw, speed_raw):
        fuel_raw = max(0.0, min(self.TANK_CAPACITY, float(fuel_raw)))
        speed_raw = max(0.0, float(speed_raw))
        previous_raw = self.raw_buffer[-1]['fuel'] if self.raw_buffer else fuel_raw
        baseline = self._baseline()
        raw_delta = fuel_raw - previous_raw
        self.raw_buffer.append({'fuel': fuel_raw, 'speed': speed_raw})

        accepted_event = baseline is None
        rejected_spike = False
        corroborated = False
        ramp_corrected_value = None
        self._retroactive_point = None

        if baseline is not None:
            transition_from_event = speed_raw > 5.0 and (
                self._refuel_ramp or self._post_refuel_hold_ticks > 0
            )
            if speed_raw > 5.0:
                self._refuel_ramp = False
                self._theft_ramp = False
                self._post_refuel_hold_ticks = max(0, self._post_refuel_hold_ticks - 1)
            ramp_expected = None
            if (speed_raw <= 5.0 and self._refuel_ramp
                    and len(self._refuel_increments) >= self.RAMP_MIN_STEPS):
                ramp_expected = float(np.median(list(self._refuel_increments)))
            ramp_deviation = (
                abs(raw_delta - ramp_expected) if ramp_expected is not None else 0.0
            )
            ramp_anomaly = (
                ramp_expected is not None
                and raw_delta >= self.EVENT_DELTA_MIN
                and ramp_deviation >= max(
                    self.RAMP_DEVIATION_MIN,
                    abs(ramp_expected) * self.RAMP_DEVIATION_RATIO
                )
            )
            if ramp_anomaly:
                ramp_corrected_value = max(
                    0.0,
                    min(self.TANK_CAPACITY, baseline + ramp_expected)
                )
                self._refuel_increments.append(ramp_expected)
                self._pending_jump = None
                rejected_spike = True
            elif (speed_raw <= 5.0 and self._refuel_ramp
                  and self.EVENT_DELTA_MIN <= raw_delta <= self.MAX_REFUEL_STEP):
                accepted_event = True
                self._refuel_increments.append(raw_delta)
                self._pending_jump = None
            elif (speed_raw <= 5.0 and self._theft_ramp
                  and -self.MAX_JUMP <= raw_delta <= -self.EVENT_DELTA_MIN):
                accepted_event = True
                self._pending_jump = None
            elif (speed_raw <= 5.0 and raw_delta >= self.REFUEL_START_JUMP
                    and self._refuel_ramp and raw_delta > self.MAX_REFUEL_STEP):
                self._pending_jump = {'raw': fuel_raw, 'delta': raw_delta, 'kind': 'spike'}
                rejected_spike = True
            elif self._is_stationary_event(raw_delta, speed_raw) and self._pending_jump is None:
                self._pending_jump = {
                    'raw': fuel_raw,
                    'delta': raw_delta,
                    'kind': 'refuel'
                }
                rejected_spike = True
            elif self._pending_jump is not None:
                pending = self._pending_jump
                follow_delta = fuel_raw - pending['raw']
                same_direction = follow_delta * pending['delta'] > 0
                between_baseline_and_pending = (
                    min(baseline, pending['raw']) <= fuel_raw <= max(baseline, pending['raw'])
                )
                if (transition_from_event and between_baseline_and_pending
                        and abs(fuel_raw - baseline) >= self.EVENT_DELTA_MIN):
                    accepted_event = True
                    corroborated = True
                    self._pending_jump = None
                    self._refuel_ramp = False
                elif (pending.get('kind') == 'transition' and transition_from_event
                      and abs(follow_delta) <= self.EVENT_DELTA_MIN):
                    accepted_event = True
                    corroborated = True
                    self._pending_jump = None
                elif pending.get('kind') == 'spike' and same_direction:
                    # A repeated jump is only allowed to become a real level
                    # change while stationary. During driving, fixed spikes
                    # must never be promoted to refuel/theft events.
                    if speed_raw <= 5.0 and abs(follow_delta) >= self.EVENT_DELTA_MIN:
                        accepted_event = True
                        corroborated = True
                        self._retroactive_point = pending['raw']
                        if pending.get('delta', 0) < 0:
                            self._theft_ramp = True
                        self._pending_jump = None
                    else:
                        self._pending_jump = {
                            'raw': fuel_raw,
                            'delta': follow_delta,
                            'kind': 'spike'
                        }
                        rejected_spike = True
                elif same_direction and abs(follow_delta) >= self.EVENT_DELTA_MIN:
                    if speed_raw <= 5.0:
                        accepted_event = True
                        corroborated = True
                        self._retroactive_point = pending['raw']
                        if pending.get('kind') == 'refuel':
                            self._refuel_ramp = True
                            self._refuel_increments.append(pending['delta'])
                        elif pending.get('delta', 0) < 0:
                            self._theft_ramp = True
                        self._pending_jump = None
                    else:
                        self._pending_jump = {
                            'raw': fuel_raw,
                            'delta': follow_delta,
                            'kind': 'spike'
                        }
                        rejected_spike = True
                elif same_direction:
                    self._pending_jump = None
                elif not same_direction:
                    self._pending_jump = None
                    crosses_baseline = (
                        pending['delta'] > 0 and fuel_raw < baseline - self.EVENT_DELTA_MIN
                    ) or (
                        pending['delta'] < 0 and fuel_raw > baseline + self.EVENT_DELTA_MIN
                    )
                    if (crosses_baseline and speed_raw <= 5.0
                            and pending.get('kind') != 'spike'):
                        accepted_event = True
                        self._refuel_ramp = False
                    elif pending.get('kind') == 'spike':
                        rejected_spike = True
                elif abs(fuel_raw - baseline) <= self.EVENT_DELTA_MIN:
                    self._pending_jump = None
                else:
                    rejected_spike = True
            elif transition_from_event and abs(raw_delta) <= self.MAX_REFUEL_STEP:
                if abs(fuel_raw - baseline) <= self.POST_REFUEL_HANDOFF_TOLERANCE:
                    accepted_event = True
                else:
                    self._pending_jump = {
                        'raw': fuel_raw,
                        'delta': fuel_raw - baseline,
                        'kind': 'spike'
                    }
                    rejected_spike = True
                self._refuel_ramp = False
            elif self._is_pending_spike(fuel_raw - baseline, speed_raw):
                self._pending_jump = {
                    'raw': fuel_raw,
                    'delta': fuel_raw - baseline,
                        'kind': 'transition' if transition_from_event else 'spike'
                }
                rejected_spike = True

        if accepted_event:
            fuel_filtered = fuel_raw
        elif ramp_corrected_value is not None:
            fuel_filtered = ramp_corrected_value
        elif rejected_spike:
            fuel_filtered = baseline
        else:
            fuel_filtered = fuel_raw if baseline is None else (
                self.EMA_ALPHA * fuel_raw + (1.0 - self.EMA_ALPHA) * baseline
            )

        fuel_filtered = max(0.0, min(self.TANK_CAPACITY, float(fuel_filtered)))
        if self._refuel_ramp and speed_raw <= 5.0 and (accepted_event or ramp_corrected_value is not None):
            self._refuel_level = fuel_filtered
        self._last_flag = {
            'flagged': rejected_spike,
            'persistent': accepted_event or corroborated,
            'classification': 'event' if accepted_event else ('spike' if rejected_spike else 'noise')
        }
        self.filtered_history.append(fuel_filtered)
        return fuel_filtered

    def filter_speed(self, speed_raw):
        self.speed_buffer.append(speed_raw)
        if len(self.speed_buffer) >= 3:
            return float(np.median(list(self.speed_buffer)))
        return speed_raw

    def reset(self, soft=False):
        if soft:
            keep = list(self.raw_buffer)[-self.RAW_HISTORY_LEN:]
            self.raw_buffer.clear()
            self.raw_buffer.extend(keep)
            self._last_flag = None
            self._pending_jump = None
            self._refuel_ramp = False
            self._theft_ramp = False
            self._refuel_increments.clear()
            self._refuel_level = float(self.filtered_history[-1]) if self.filtered_history else None
            self._post_refuel_hold_ticks = self.POST_REFUEL_HOLD_TICKS if self._refuel_level is not None else 0
            if self.filtered_history:
                self.kalman.anchor(float(self.filtered_history[-1]), p=1.0)
            else:
                self.kalman.p = max(self.kalman.p, 1.0)
        else:
            self.kalman.reset()
            self.raw_buffer.clear()
            self.speed_buffer.clear()
            self.filtered_history.clear()
            self.tracker_status = 'normal'
            self.event_type = None
            self._last_flag = None
            self._pending_jump = None
            self._refuel_ramp = False
            self._theft_ramp = False
            self._refuel_increments.clear()
            self._refuel_level = None
            self._post_refuel_hold_ticks = 0


class AdaptiveDeltaDetector:
    """
    V3: Tách gate theo 3 vùng speed + cumulative delta để bắt Theft ~4L/tick.
    - speed>15: GATE=18L, cần 3 tick (tránh FP từ spike lớn khi Driving)
    - 5<speed<=15: GATE=10L, cần 2 tick
    - speed<=5: GATE=4L, confirm khi 2 tick HOẶC cumulative>12L
    - delta>=15L bất kể speed: refuel ngay (1 tick)
    """
    GATE_FAST = 15.0
    GATE_DRIVING = 18.0
    GATE_TRANS = 10.0
    GATE_IDLE = 4.0

    def __init__(self):
        self.raw_history = deque(maxlen=8)
        self._ema = None
        self._dir = None
        self._ticks = 0
        self._cumul = 0.0

    def update(self, raw_fuel, speed, filtered_fuel=None):
        signal_value = raw_fuel if filtered_fuel is None else filtered_fuel
        previous_value = self.raw_history[-1] if self.raw_history else signal_value
        self._ema = signal_value
        self.raw_history.append(signal_value)
        if len(self.raw_history) < 2:
            return None, 'stable'

        delta = signal_value - previous_value

        # IMPORTANT: a single large spike must not act as a confirmed refuel/theft event.
        # Only a persistent multi-tick trend is considered strong enough to trigger an event.
        # This prevents the 1-tick GATE_FAST branch from bypassing all persistence logic.
        direction = 'up' if delta > 0 else 'down'

        if speed > 15.0:
            if abs(delta) >= self.GATE_DRIVING:
                self._accum(direction, abs(delta))
                if self._ticks >= 3:
                    return abs(delta), 'refuel' if direction == 'up' else 'theft'
                return abs(delta), 'pending'
            self._reset()
            return abs(delta), 'driving_noise'

        elif speed > 5.0:
            if abs(delta) >= self.GATE_TRANS:
                self._accum(direction, abs(delta))
                if self._ticks >= 2:
                    return abs(delta), 'refuel' if direction == 'up' else 'theft'
                return abs(delta), 'pending'
            self._reset()
            return abs(delta), 'stable'

        else:
            if abs(delta) >= self.GATE_IDLE:
                self._accum(direction, abs(delta))
                if self._ticks >= 2 or self._cumul >= 12.0:
                    return abs(delta), 'refuel' if direction == 'up' else 'theft'
                return abs(delta), 'pending'
            elif abs(delta) >= 1.5 and self._dir == direction:
                self._ticks += 0.5
                self._cumul += abs(delta)
                if self._cumul >= 12.0:
                    return abs(delta), 'refuel' if direction == 'up' else 'theft'
                return abs(delta), 'pending'
            self._reset()
            return abs(delta), 'stable'

    def _accum(self, direction, mag):
        if direction == self._dir:
            self._ticks += 1
            self._cumul += mag
        else:
            self._dir = direction
            self._ticks = 1
            self._cumul = mag

    def anchor_ema(self, value):
        self._ema = value

    def _reset(self):
        self._dir = None
        self._ticks = 0
        self._cumul = 0.0

    def _reset_candidate(self):
        self._reset()

    def reset(self):
        self._reset()
        self.raw_history.clear()


class EventTracker:
    def __init__(self):
        self.last_prediction = None
        self.last_confidence = 0.0
        self.confidence_history = []
        self._processed_buffer_ref = None
        self._raw_buffer_ref = None
        self.reset()
        self.last_finished_event = None
        self._last_event_end_ts = None
        self._post_refuel_flag = False
        self._post_refuel_ticks = 0
        self._current_window = 8
        self._pending_switch = None
        # IMPROVEMENT #3: Smart post-refuel skip tracking
        self._fuel_deltas_recent = deque(maxlen=3)

    def reset(self):
        self.status = 'normal'
        self.candidate_label = None
        self.candidate_start = None
        self.counter = 0
        self.best_conf = 0.0
        self.event_record = None
        if len(self.confidence_history) > 200:
            self.confidence_history = self.confidence_history[-200:]
        self.stage_history = []
        self._just_confirmed = False
        self.confirmed_ticks = 0
        self.mismatch_streak = 0
        self._post_refuel_flag = False
        self._pending_switch = None

    @property
    def point_status(self):
        if self.status == 'candidate':
            return 'thinking'
        if self.status in ('confirmed', 'finished'):
            return 'confirmed'
        return 'normal'

    @property
    def display_stage(self):
        if self.status == 'confirmed':
            return 'confirmed' if self._just_confirmed else 'monitoring'
        return self.status

    @property
    def active_event_label(self):
        if self.status == 'candidate':
            return self.candidate_label
        if self.status in ('confirmed', 'finished') and self.event_record:
            return self.event_record.get('event')
        return None

    def _push_stage(self, stage, timestamp):
        self.stage_history.append({'stage': stage, 'timestamp': _ts_to_iso(timestamp)})

    def _get_confirm_threshold(self):
        if self.candidate_label == 'Idle':
            return 0.5
        w = max(2, min(8, self._current_window))
        reduction = (8 - w) * 0.02
        return max(0.55, CONFIRM_THRESH - reduction)

    def get_active_summary(self, full_df=None):
        if self.status != 'normal':
            event_type = self.candidate_label if self.candidate_label else (
                self.event_record.get('event') if self.event_record else None)
            fuel_start_val = None
            fuel_current_val = None
            delta_fuel = 0.0
            duration_s = 0
            fuel_status = 'Stable'

            if full_df is not None and len(full_df) > 0:
                fuel_current_val = float(full_df.iloc[-1]['fuel'])
                if self.event_record is not None:
                    if self.event_record.get('fuel_before'):
                        fuel_start_val = self.event_record['fuel_before']
                    else:
                        start_ts = self.candidate_start if self.candidate_start else (
                            self.event_record.get('start_time') if self.event_record else None)
                        if start_ts:
                            df = full_df.sort_values('timestamp')
                            start_idx = df['timestamp'].searchsorted(start_ts)
                            if start_idx < len(df):
                                fuel_start_val = float(df.iloc[start_idx]['fuel'])
                if fuel_start_val is not None:
                    delta_fuel = fuel_current_val - fuel_start_val
                    if delta_fuel > 0.5:
                        fuel_status = 'Refueling'
                    elif delta_fuel < 0:
                        fuel_status = 'Theft' if event_type == 'Theft' else 'Consuming'
                    else:
                        fuel_status = 'Stable'

                if self.status == 'finished':
                    if self.event_record and self.event_record.get('duration_s'):
                        duration_s = self.event_record['duration_s']
                else:
                    start_ts = None
                    if self.event_record and self.event_record.get('start_time'):
                        start_ts = self.event_record['start_time']
                    elif self.candidate_start:
                        start_ts = self.candidate_start
                    if start_ts and full_df is not None:
                        duration_s = int((full_df.iloc[-1]['timestamp'] - start_ts).total_seconds())

            return {
                'event_type': event_type,
                'stage': self.display_stage,
                'start_time': _ts_to_iso(self.candidate_start) if self.candidate_start else None,
                'confidence': self.best_conf,
                'stage_history': list(self.stage_history) if self.stage_history else [
                    {'stage': self.display_stage, 'timestamp': _ts_to_iso(self.candidate_start)}],
                'fuel_start': fuel_start_val,
                'fuel_current': fuel_current_val,
                'delta_fuel': round(delta_fuel, 2),
                'fuel_status': fuel_status,
                'duration_s': duration_s
            }

        if full_df is not None and len(full_df) > 0:
            current_ts = full_df.iloc[-1]['timestamp']
            fuel_current_val = float(full_df.iloc[-1]['fuel'])
            self.stage_history = [{'stage': 'normal', 'timestamp': _ts_to_iso(current_ts)}]
            return {
                'event_type': self.last_prediction or 'Driving',
                'stage': 'normal',
                'start_time': _ts_to_iso(current_ts),
                'confidence': self.last_confidence if self.last_prediction else 1.0,
                'stage_history': self.stage_history,
                'fuel_start': None,
                'fuel_current': fuel_current_val,
                'delta_fuel': 0.0,
                'fuel_status': 'Stable',
                'duration_s': 0
            }
        return None

    def pop_finished_event(self):
        ev = self.last_finished_event
        self.last_finished_event = None
        return ev

    def retroactively_fix_point_status(self, buffer_df, processed_buffer):
        if self.status != 'confirmed' or not self.event_record:
            return
        start_ts = self.event_record['start_time']
        event_label = self.event_record['event']
        for record in processed_buffer:
            if (record['Timestamp'] >= start_ts
                    and record.get('PointStatus') == 'thinking'
                    and record.get('Prediction') == event_label):
                record['PointStatus'] = 'confirmed'

    def _retroactively_clear_false_candidate(self, buffer_df):
        if not self._processed_buffer_ref or not self.candidate_start:
            return
        start_ts = self.candidate_start
        bad_label = self.candidate_label
        for record in reversed(self._processed_buffer_ref):
            if record['Timestamp'] < start_ts:
                break
            if record.get('PointStatus') == 'thinking' and record.get('Prediction') == bad_label:
                record['PointStatus'] = 'normal'

    def _robust_delta(self, buffer_df, lookback=4):
        if buffer_df is None or len(buffer_df) < 3:
            return 0.0
        df = buffer_df.sort_values('timestamp').reset_index(drop=True)
        if len(df) <= lookback:
            if len(df) >= 2:
                return float(df.iloc[-1]['fuel']) - float(df.iloc[-2]['fuel'])
            return 0.0
        baseline_points = df.iloc[-(lookback + 1):-1]['fuel'].astype(float)
        baseline = float(baseline_points.median()) if len(baseline_points) else float(df.iloc[-2]['fuel'])
        current = float(df.iloc[-1]['fuel'])
        return current - baseline

    def _should_finish_refuel_transition(self, label_name, conf, buffer_df):
        if label_name != 'Driving' or conf < 0.65 or buffer_df is None or len(buffer_df) < 3:
            return False
        df = buffer_df.sort_values('timestamp').reset_index(drop=True)
        current_speed = float(df.iloc[-1]['speed'])
        if current_speed <= 5.0:
            return False
        recent = df.iloc[-3:]
        fuels = recent['fuel'].astype(float).tolist()
        delta_last = fuels[-1] - fuels[-2]
        delta_net = fuels[-1] - fuels[0]
        if abs(delta_last) <= 1.5 and abs(delta_net) <= 3.0:
            return True
        return False

    def _has_credible_theft_trend(self, buffer_df):
        if buffer_df is None or len(buffer_df) < 3:
            return False
        df = buffer_df.sort_values('timestamp').reset_index(drop=True)
        current_speed = float(df.iloc[-1]['speed'])
        if current_speed > 5.0:
            return False
        recent = df.iloc[-3:]
        fuels = recent['fuel'].astype(float).tolist()
        if len(fuels) < 2:
            return False
        net_drop = fuels[-1] - fuels[0]
        last_drop = fuels[-1] - fuels[-2]
        return (net_drop <= -THEFT_MIN_DELTA) or (last_drop <= -THEFT_MIN_DELTA * 0.75)

    def _pending_switch_is_valid(self, label_name, buffer_df):
        if buffer_df is None or len(buffer_df) < 2:
            return False
        if label_name == 'Driving':
            df = buffer_df.sort_values('timestamp').reset_index(drop=True)
            recent_speeds = df.iloc[-min(3, len(df)):]['speed'].astype(float).tolist()
            if not recent_speeds:
                return False
            return float(df.iloc[-1]['speed']) > DRIVING_SPEED_GATE and float(np.median(recent_speeds)) > DRIVING_SPEED_GATE * 0.75
        robust_delta = self._robust_delta(buffer_df, lookback=4)
        if label_name == 'Refuel':
            return robust_delta >= max(REFUEL_MIN_DELTA, 3.0)
        if label_name == 'Theft':
            return robust_delta <= -max(THEFT_MIN_DELTA, 3.0)
        return False

    def _should_permit_idle_bypass(self, label_name, buffer_df):
        if label_name not in ('Refuel', 'Theft') or buffer_df is None or len(buffer_df) < 4:
            return False
        robust_delta = self._robust_delta(buffer_df, lookback=4)
        if label_name == 'Refuel':
            return robust_delta >= max(REFUEL_MIN_DELTA, 3.0)
        if label_name == 'Theft':
            return robust_delta <= -max(THEFT_MIN_DELTA, 3.0)
        return False

    def _maybe_reopen_candidate(self, label_name, conf, timestamp, buffer_df=None):
        if self.status != 'normal':
            return
        if conf < CANDIDATE_THRESH:
            return
        if label_name in ('Refuel', 'Theft') and buffer_df is not None and len(buffer_df) >= 2:
            robust_delta = self._robust_delta(buffer_df, lookback=4)
            if label_name == 'Refuel' and robust_delta < REFUEL_MIN_DELTA:
                return
            if label_name == 'Theft' and robust_delta > -THEFT_MIN_DELTA:
                return
        self.status = 'candidate'
        self.candidate_label = label_name
        self.candidate_start = timestamp
        self.counter = 1
        self.best_conf = conf
        self.mismatch_streak = 0
        self.stage_history = [{'stage': 'normal', 'timestamp': _ts_to_iso(timestamp)}]
        self._push_stage('candidate', timestamp)

    def update(self, prob, buffer_df, timestamp):
        label_id = int(np.argmax(prob))
        conf = float(prob[label_id])
        label_name = LABELS[label_id]
        self.last_prediction = label_name
        self.last_confidence = conf
        self.confidence_history.append((timestamp, label_name, conf))
        if len(self.confidence_history) > 200:
            self.confidence_history.pop(0)

        if self._pending_switch is not None:
            if label_name == self._pending_switch['label'] and self._pending_switch_is_valid(label_name, buffer_df):
                self.reset()
                self._maybe_reopen_candidate(label_name, conf, timestamp, buffer_df)
                return
            self._pending_switch = None

        if self.status == 'normal':
            self._maybe_reopen_candidate(label_name, conf, timestamp, buffer_df)

        elif self.status == 'candidate':
            if label_name != self.candidate_label:
                if (self.candidate_label == 'Idle'
                        and label_name == 'Driving'
                        and buffer_df is not None and len(buffer_df) >= 2):
                    recent_speeds = [float(x) for x in buffer_df.iloc[-min(3, len(buffer_df)):]['speed']]
                    if (recent_speeds and float(buffer_df.iloc[-1]['speed']) > DRIVING_SPEED_GATE
                            and float(np.median(recent_speeds)) > DRIVING_SPEED_GATE * 0.75):
                        self._pending_switch = {'label': label_name, 'ts': timestamp}
                        return

                if (self.candidate_label == 'Idle'
                        and label_name in ('Refuel', 'Theft')
                        and buffer_df is not None and len(buffer_df) >= 4):
                    if self._should_permit_idle_bypass(label_name, buffer_df):
                        self._pending_switch = {'label': label_name, 'ts': timestamp}
                        return

                self.mismatch_streak += 1
                tolerance = 2 if self.candidate_label == 'Idle' else (
                    FUEL_EVENT_MISMATCH_TOLERANCE
                    if self.candidate_label in ('Refuel', 'Theft')
                    else CANDIDATE_MISMATCH_TOLERANCE)
                if self.candidate_label == 'Idle':
                    current_speed = float(buffer_df.iloc[-1]['speed'])
                    recent_fuel = [buffer_df.iloc[-i-1]['fuel']
                                   for i in range(min(3, len(buffer_df)))] if len(buffer_df) >= 3 else [0, 0, 0]
                    fuel_slope = abs(recent_fuel[0] - recent_fuel[-1]) / max(1, len(recent_fuel) - 1)
                    if current_speed < 1.0 and fuel_slope < IDLE_FUEL_SLOPE_RESET:
                        self.mismatch_streak = 0
                if self.candidate_label == 'Theft' and self._has_credible_theft_trend(buffer_df):
                    self.mismatch_streak = 0
                    return
                if self.mismatch_streak > tolerance:
                    self._retroactively_clear_false_candidate(buffer_df)
                    self.reset()
                    self._maybe_reopen_candidate(label_name, conf, timestamp, buffer_df)
                return
            else:
                self.mismatch_streak = 0

            if conf >= CANDIDATE_THRESH:
                self.counter += 1
                self.best_conf = max(self.best_conf, conf)
                effective_confirm_thresh = self._get_confirm_threshold()
                if self.counter >= CONFIRM_COUNT and conf >= effective_confirm_thresh:
                    self.status = 'confirmed'
                    self._just_confirmed = True
                    self.confirmed_ticks = 0
                    self.mismatch_streak = 0
                    self.event_record = {
                        'event': self.candidate_label,
                        'start_time': self.candidate_start,
                        'confidence': self.best_conf,
                    }
                    self._backtrack_start(buffer_df)
                    self._push_stage('confirmed', timestamp)
                    summary = self.get_active_summary(buffer_df)
                    _append_toast_log('START', self.car_id, {
                        'event': self.candidate_label,
                        'start_time': self.event_record['start_time'],
                        'fuel_before': summary.get('fuel_start'),
                        'fuel_after': summary.get('fuel_current'),
                        'fuel_change': summary.get('delta_fuel'),
                        'duration_s': summary.get('duration_s')
                    })
            else:
                self.mismatch_streak += 1
                tolerance = (FUEL_EVENT_MISMATCH_TOLERANCE
                             if self.candidate_label in ('Refuel', 'Theft')
                             else CANDIDATE_MISMATCH_TOLERANCE)
                if self.mismatch_streak > tolerance:
                    self._retroactively_clear_false_candidate(buffer_df)
                    self.reset()

        elif self.status == 'confirmed':
            self.confirmed_ticks += 1
            if self._just_confirmed:
                self._just_confirmed = False
                self._push_stage('monitoring', timestamp)

            self.event_record['fuel_current'] = float(buffer_df.iloc[-1]['fuel'])
            try:
                df = buffer_df.sort_values('timestamp')
                start_idx = df['timestamp'].searchsorted(self.event_record['start_time'])
                if start_idx < len(df):
                    window_fuel = df.iloc[start_idx:]['fuel']
                    if len(window_fuel) > 0:
                        self.event_record['min_fuel'] = round(float(window_fuel.min()), 2)
            except Exception as e:
                logger.error(f"Active Min Fuel calc error: {e}")

            current_speed = float(buffer_df.iloc[-1]['speed'])
            event_type = self.event_record.get('event')
            if event_type in ['Refuel', 'Theft'] and current_speed > 10.0:
                self._finish_event(buffer_df, timestamp)
                self._maybe_reopen_candidate(label_name, conf, timestamp, buffer_df)
                return

            effective_finish = FINISH_THRESH
            if event_type in ('Refuel', 'Theft'):
                w = max(2, min(8, self._current_window))
                effective_finish = max(0.35, FINISH_THRESH - (8 - w) * 0.015)

            transition_to_driving = False
            if event_type in ('Refuel', 'Theft') and self._should_finish_refuel_transition(label_name, conf, buffer_df):
                transition_to_driving = True

            if event_type == 'Theft' and self._has_credible_theft_trend(buffer_df):
                self.best_conf = max(self.best_conf, conf)
                self.event_record['confidence'] = self.best_conf
                return

            if transition_to_driving or conf < effective_finish or label_name != self.event_record['event']:
                self._finish_event(buffer_df, timestamp)
                self._maybe_reopen_candidate(label_name, conf, timestamp, buffer_df)
            else:
                self.best_conf = max(self.best_conf, conf)
                self.event_record['confidence'] = self.best_conf

    def _is_weak_event(self, event_type, fuel_change, duration_s):
        if event_type not in ('Driving', 'Idle'):
            return False

        min_delta = DRIVING_MIN_FUEL_DELTA if event_type == 'Driving' else IDLE_MIN_FUEL_DELTA
        if abs(float(fuel_change)) >= min_delta:
            return False
        if int(duration_s) >= WEAK_EVENT_DURATION_GRACE_S:
            return False
        return True

    def _discard_weak_event(self, buffer_df):
        if not self.event_record:
            return

        start_ts = self.event_record.get('start_time')
        event_label = self.event_record.get('event')
        if not start_ts or not event_label:
            self.reset()
            return

        if self._processed_buffer_ref is None:
            self.reset()
            return

        for record in reversed(self._processed_buffer_ref):
            if record.get('Timestamp') is None:
                continue
            if record['Timestamp'] < start_ts:
                break
            if record.get('Prediction') == event_label and record.get('PointStatus') in ('thinking', 'confirmed', 'monitoring'):
                record['PointStatus'] = 'normal'

        logger.info(
            f"🧹 [Car {self.car_id}] Weak event discarded: {event_label} | "
            f"fuel_change={self.event_record.get('fuel_change', 0)} | "
            f"duration_s={self.event_record.get('duration_s', 0)}"
        )
        self.reset()

    def _finish_event(self, buffer_df, end_time):
        self.status = 'finished'
        self._backtrack_end(buffer_df, end_time)
        self._push_stage('finished', end_time)

        df = buffer_df.sort_values('timestamp').reset_index(drop=True)
        start_ts = self.event_record['start_time']
        end_ts = self.event_record['end_time']
        min_fuel = None
        try:
            start_idx = int(df['timestamp'].searchsorted(start_ts))
            end_idx = int(df['timestamp'].searchsorted(end_ts))
            start_idx = max(0, min(start_idx, len(df) - 1))
            end_idx = max(0, min(end_idx, len(df) - 1))
            if start_idx < len(df) and end_idx < len(df):
                self.event_record['fuel_before'] = round(float(df.iloc[start_idx]['fuel']), 2)
                self.event_record['fuel_after'] = round(float(df.iloc[end_idx]['fuel']), 2)
                self.event_record['fuel_change'] = round(
                    self.event_record['fuel_after'] - self.event_record['fuel_before'], 2)
                lo, hi = sorted([start_idx, end_idx])
                window_fuel = df.iloc[lo:hi + 1]['fuel']
                if len(window_fuel):
                    min_fuel = round(float(window_fuel.min()), 2)
        except Exception as e:
            logger.error(f"Fuel change error: {e}")

        self.event_record['min_fuel'] = min_fuel
        self.event_record['fuel_added'] = self.event_record.get('fuel_change', 0)
        self.event_record['max_confidence'] = self.best_conf
        self.event_record['duration_s'] = int(
            (self.event_record['end_time'] - self.event_record['start_time']).total_seconds())
        self.event_record['stage_history'] = list(self.stage_history)

        if self._is_weak_event(
            self.event_record.get('event'),
            self.event_record.get('fuel_change', 0),
            self.event_record.get('duration_s', 0)
        ):
            self._discard_weak_event(buffer_df)
            return

        was_refuel = (self.event_record['event'] == 'Refuel')
        self.last_finished_event = dict(self.event_record)
        self._last_event_end_ts = self.event_record.get('end_time')
        _append_toast_log('END', self.car_id, self.event_record)
        self.reset()

        if was_refuel:
            self._post_refuel_flag = True
            self._post_refuel_ticks = 3

    def _backtrack_start(self, buffer_df):
        threshold = BACKTRACK_FUEL_THRESHOLD
        event_label = self.event_record['event']
        df = buffer_df.sort_values('timestamp').reset_index(drop=True)
        idx = df['timestamp'].searchsorted(self.event_record['start_time'])
        if idx >= len(df):
            idx = len(df) - 1

        last_end_ts = None
        if self._last_event_end_ts is not None:
            last_end_ts = pd.Timestamp(self._last_event_end_ts)

        if event_label in ('Refuel', 'Theft') and idx > 0:
            min_delta = REFUEL_MIN_DELTA if event_label == 'Refuel' else THEFT_MIN_DELTA
            prev_fuel = float(df.iloc[idx - 1]['fuel'])
            cur_fuel = float(df.iloc[idx]['fuel'])
            delta = cur_fuel - prev_fuel
            cond = delta >= min_delta if event_label == 'Refuel' else delta <= -min_delta
            if cond:
                self.event_record['start_time'] = df.iloc[idx - 1]['timestamp']
                self.event_record['boundary_point'] = True
                if last_end_ts is not None:
                    if pd.Timestamp(self.event_record['start_time']) < last_end_ts:
                        self.event_record['start_time'] = last_end_ts
                return

        label_by_ts = {}
        for ts, lbl, _ in self.confidence_history:
            label_by_ts[pd.Timestamp(ts)] = lbl

        lower_bound = max(0, idx - BACKTRACK_MAX_LOOKBACK)
        new_start_idx = idx
        for i in range(idx, lower_bound, -1):
            cur_ts = pd.Timestamp(df.iloc[i]['timestamp'])
            cur_label = label_by_ts.get(cur_ts)
            if cur_label is not None and cur_label != event_label:
                break
            new_start_idx = i
            if i > 0:
                prev_ts = pd.Timestamp(df.iloc[i - 1]['timestamp'])
                prev_label = label_by_ts.get(prev_ts)
                if prev_label is not None and prev_label != event_label:
                    break
                if abs(df.iloc[i]['fuel'] - df.iloc[i - 1]['fuel']) >= threshold:
                    break

        if last_end_ts is not None:
            last_idx = int(df['timestamp'].searchsorted(last_end_ts))
            last_idx = max(0, min(last_idx, len(df) - 1))
            if last_idx > new_start_idx:
                new_start_idx = last_idx

        self.event_record['start_time'] = df.iloc[new_start_idx]['timestamp']
        if last_end_ts is not None and pd.Timestamp(self.event_record['start_time']) < last_end_ts:
            self.event_record['start_time'] = last_end_ts

        if self.event_record.get('start_time') is not None and last_end_ts is not None:
            start_ts = pd.Timestamp(self.event_record['start_time'])
            if start_ts < last_end_ts:
                self.event_record['start_time'] = last_end_ts

    def _backtrack_end(self, buffer_df, end_time):
        df = buffer_df.sort_values('timestamp').reset_index(drop=True)
        idx = df['timestamp'].searchsorted(end_time)
        if idx >= len(df):
            idx = len(df) - 1
        event_label = self.event_record.get('event')

        if event_label in ('Driving', 'Idle') and idx > 0:
            prev_fuel = float(df.iloc[idx - 1]['fuel'])
            cur_fuel = float(df.iloc[idx]['fuel'])
            jump = cur_fuel - prev_fuel
            transition_threshold = max(REFUEL_MIN_DELTA, THEFT_MIN_DELTA * 2)
            if abs(jump) >= transition_threshold:
                self.event_record['end_time'] = df.iloc[idx - 1]['timestamp']
                return

        lower_bound = max(0, idx - 5)
        new_end_idx = idx
        for i in range(idx, lower_bound, -1):
            cur_speed = df.iloc[i]['speed']
            prev_speed = df.iloc[i - 1]['speed'] if i > 0 else cur_speed
            if event_label == 'Driving' and prev_speed > 8.0 and cur_speed < 2.0:
                new_end_idx = i
                break
            if event_label == 'Idle' and prev_speed < 2.0 and cur_speed > 8.0:
                new_end_idx = i
                break
        self.event_record['end_time'] = df.iloc[new_end_idx]['timestamp']


class AdaptiveWindowManager:
    def __init__(self, ranges=None, idle_steps=None, stability_std=STABILITY_STD_THRESHOLD):
        self.ranges = ranges or WINDOW_RANGES
        self.idle_steps = idle_steps or IDLE_STEPS
        self.stability_std = stability_std

    def get_window_size(self, tracker: EventTracker, data_buffer=None) -> int:
        status = tracker.status

        if status == 'candidate':
            if tracker.candidate_label == 'Idle':
                return self.ranges['Idle'][0]
            if tracker.candidate_label in ('Refuel', 'Theft'):
                return self.ranges['Refuel'][0]
            if tracker.candidate_label == 'Driving' and data_buffer:
                recent_speeds = [float(point.get('speed', 0))
                                 for point in list(data_buffer)[-3:]]
                if recent_speeds and recent_speeds[-1] > DRIVING_SPEED_GATE:
                    return self.ranges['Driving'][0]
            return self.ranges['candidate'][0]

        if status in ('confirmed', 'finished'):
            event = tracker.active_event_label
            if event in ('Refuel', 'Theft') and data_buffer:
                recent_speeds = [float(point.get('speed', 0))
                                 for point in list(data_buffer)[-3:]]
                if recent_speeds and recent_speeds[-1] > DRIVING_SPEED_GATE:
                    return self.ranges['Driving'][0]
            lo, hi = self.ranges.get(event, self.ranges['normal'])
            if event in ('Refuel', 'Theft'):
                return lo
            if event == 'Idle':
                return self._idle_step(tracker.confirmed_ticks)
            return self._stability_adjusted(lo, hi, tracker.confidence_history)

        if getattr(tracker, '_post_refuel_ticks', 0) > 0 or getattr(tracker, '_post_refuel_flag', False):
            if data_buffer:
                recent_speeds = [float(point.get('speed', 0))
                                 for point in list(data_buffer)[-3:]]
                if recent_speeds and recent_speeds[-1] > DRIVING_SPEED_GATE:
                    return self.ranges['Driving'][0]
            return 3

        if data_buffer and len(data_buffer) >= 2:
            recent_speeds = [p['speed'] for p in list(data_buffer)[-3:]]
            avg_speed = float(np.mean(recent_speeds))
            if avg_speed < 1.0:
                return 2
            if avg_speed < 5.0:
                return 3

        sustained = self._sustained_change(data_buffer)
        if sustained > 20.0:
            return 4
        if sustained > 10.0:
            return 5
        return self.ranges['normal'][1]

    def _sustained_change(self, data_buffer):
        if not data_buffer or len(data_buffer) < 4:
            return 0.0
        pts = list(data_buffer)[-4:]
        fuels = [p['fuel'] for p in pts]
        recent = float(np.median(fuels[-2:]))
        prior = float(np.median(fuels[:2]))
        speeds = [p['speed'] for p in pts]
        avg_speed = float(np.mean(speeds))
        if avg_speed > 5.0 and (recent - prior) > 10.0:
            return 0.0
        return abs(recent - prior)

    def _stability_adjusted(self, lo, hi, confidence_history):
        if hi <= lo or len(confidence_history) < 2:
            return lo
        recent = [c for _, _, c in confidence_history[-lo:]]
        if len(recent) < 2:
            return lo
        std = float(np.std(recent))
        if std <= self.stability_std:
            return lo
        ratio = min(std / 0.3, 1.0)
        window = lo + int(round((hi - lo) * ratio))
        return max(lo, min(hi, window))

    def _idle_step(self, confirmed_ticks):
        idx = min(max(confirmed_ticks, 0), len(self.idle_steps) - 1)
        return max(3, self.idle_steps[idx])


class SegmentBuilder:
    @staticmethod
    def build(data_buffer, window_size):
        available = len(data_buffer)
        actual_window = max(1, min(window_size, available))
        window_list = list(data_buffer)[-actual_window:]
        segment_df = pd.DataFrame(window_list).sort_values('timestamp').reset_index(drop=True)
        return segment_df, actual_window


class ModelWrapper:
    def __init__(self, model_path):
        self.model = CNNGRUClassifier(
            input_dim=4, hidden_dim=128, num_layers=2, dropout=0.3,
            bidirectional=True, feature_dim=21, num_classes=4,
            fusion_type='concat', fusion_dim=256,
            cnn_channels=[64, 128], cnn_kernel_size=3
        ).to(DEVICE)
        self.model.load_state_dict(torch.load(model_path, map_location=DEVICE))
        self.model.eval()
        logger.info("✅ AI Model loaded successfully.")

    def predict(self, sequence, mask, feature):
        with torch.no_grad():
            seq_t = torch.from_numpy(sequence).float().unsqueeze(0).to(DEVICE)
            mask_t = torch.from_numpy(mask).bool().unsqueeze(0).to(DEVICE)
            feat_t = torch.from_numpy(feature).float().unsqueeze(0).to(DEVICE)
            logits = self.model(seq_t, mask_t, feat_t)
            return torch.softmax(logits, dim=1).cpu().numpy()[0]


def make_json_safe(obj):
    if isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating,)):
        return float(obj)
    elif isinstance(obj, (pd.Timestamp, datetime)):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [make_json_safe(item) for item in obj]
    return obj

def _ts_to_iso(ts):
    return ts.isoformat() if hasattr(ts, 'isoformat') else str(ts)


def _append_toast_log(log_type, car_id, data):
    try:
        entry = {
            'type': log_type, 'car_id': car_id,
            'event': data.get('event'),
            'start_time': _ts_to_iso(data.get('start_time')),
            'end_time': _ts_to_iso(data.get('end_time')) if data.get('end_time') else None,
            'fuel_start': round(data.get('fuel_before', 0), 2) if data.get('fuel_before') is not None else None,
            'fuel_current': round(data.get('fuel_after', 0), 2) if data.get('fuel_after') is not None else None,
            'delta': round(data.get('fuel_change', 0), 2) if data.get('fuel_change') is not None else None,
            'duration_s': data.get('duration_s'),
            'logged_at': datetime.now().isoformat()
        }
        cleaned_entry = {k: v for k, v in entry.items() if v is not None}
        with open(TOAST_LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(json.dumps(make_json_safe(cleaned_entry), ensure_ascii=False) + "\n")
    except Exception as e:
        logger.error(f"⚠️ Không ghi được toast log: {e}")

def _read_toast_log(limit=4, car_id=None):
    if not os.path.exists(TOAST_LOG_PATH):
        return []
    lines = []
    try:
        with open(TOAST_LOG_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                if car_id and obj.get('car_id') != car_id:
                    continue
                lines.append(obj)
        return lines[-limit:]
    except Exception as e:
        logger.error(f"⚠️ Không đọc được toast log: {e}")
        return []

def _append_event_log(record, car_id):
    try:
        entry = dict(record)
        entry['car_id'] = car_id
        entry['logged_at'] = datetime.now().isoformat()
        with open(EVENT_LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(json.dumps(make_json_safe(entry), ensure_ascii=False) + "\n")
        logger.info(f"📝 Event log saved: {car_id} | {record.get('event')} | "
                    f"{record.get('start_time')} → {record.get('end_time')}")
    except Exception as e:
        logger.error(f"⚠️ Không ghi được event log: {e}")

def _read_event_log(limit=200, car_id=None):
    if not os.path.exists(EVENT_LOG_PATH):
        return []
    lines = []
    try:
        with open(EVENT_LOG_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                if car_id and obj.get('car_id') != car_id:
                    continue
                lines.append(obj)
    except Exception as e:
        logger.error(f"⚠️ Không đọc được event log: {e}")
        return []
    return lines[-limit:]


_db_lock = threading.Lock()

def _get_db_conn():
    conn = sqlite3.connect(REALTIME_DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL;")
    return conn

def _init_db():
    with _db_lock:
        conn = _get_db_conn()
        conn.execute("""CREATE TABLE IF NOT EXISTS realtime_points (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                car_id TEXT NOT NULL, timestamp TEXT NOT NULL,
                fuel REAL, fuel_raw REAL, fuel_delta REAL,
                speed REAL, latitude REAL, longitude REAL,
                prediction TEXT, confidence REAL, point_status TEXT,
                window_size INTEGER, created_at TEXT
            )""")
        existing_cols = {row[1] for row in conn.execute("PRAGMA table_info(realtime_points)").fetchall()}
        for col, col_type in [('fuel_raw', 'REAL'), ('fuel_delta', 'REAL')]:
            if col not in existing_cols:
                conn.execute(f"ALTER TABLE realtime_points ADD COLUMN {col} {col_type}")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_realtime_points_car_ts ON realtime_points (car_id, timestamp)")
        conn.commit()
        conn.close()

def _save_point_to_db(car_id, record):
    try:
        with _db_lock:
            conn = _get_db_conn()
            conn.execute(
                """INSERT INTO realtime_points
                   (car_id, timestamp, fuel, fuel_raw, fuel_delta, speed, latitude, longitude,
                    prediction, confidence, point_status, window_size, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (car_id, _ts_to_iso(record['Timestamp']),
                 float(record.get('Fuel', 0)), float(record.get('Fuel_Raw', 0)),
                 float(record.get('Fuel_Delta', 0)), float(record.get('Speed', 0)),
                 float(record.get('Latitude', 0)), float(record.get('Longitude', 0)),
                 record.get('Prediction'), float(record.get('Confidence', 0)),
                 record.get('PointStatus', 'normal'), int(record.get('WindowSize', 0)),
                 datetime.now().isoformat())
            )
            conn.commit()
            conn.close()
    except Exception as e:
        logger.error(f"⚠️ Không ghi được realtime point vào DB: {e}")

_init_db()


class CarState:
    def __init__(self, car_id, prefill_points=None):
        self.car_id = car_id
        self.data_buffer_raw = deque(maxlen=BUFFER_SIZE)
        self.data_buffer_filtered = deque(maxlen=BUFFER_SIZE)
        self.noise_filter = AdaptiveNoiseFilter()
        self.delta_detector = AdaptiveDeltaDetector()
        if prefill_points:
            for p in prefill_points:
                self.data_buffer_raw.append(p)
                self.data_buffer_filtered.append(p.copy())
        self.processed_buffer = []
        self.event_history = []
        self.tracker = EventTracker()
        self.tracker.car_id = car_id
        self.tracker._processed_buffer_ref = self.processed_buffer
        self.tracker._raw_buffer_ref = self.data_buffer_raw
        self.stationary_ticks = 0
        self.moving_ticks = 0
        self.lock = threading.Lock()

car_data_cache = {}
car_states = {}
_car_states_meta_lock = threading.Lock()
current_car = "Car 2"


def load_all_cars():
    if not os.path.exists(FUSION_CSV):
        logger.warning(f"⚠️ Không tìm thấy {FUSION_CSV}")
        return
    df = pd.read_csv(FUSION_CSV)
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    for car in df['car_id'].unique():
        car_df = df[df['car_id'] == car].sort_values('timestamp').reset_index(drop=True)
        car_data_cache[car] = car_df
    logger.info(f"✅ Đã nạp dữ liệu nền cho {len(car_data_cache)} xe: {list(car_data_cache.keys())}")

load_all_cars()

def _baseline_points_for_car(car_id, n=WINDOW_SIZE):
    df = car_data_cache.get(car_id)
    if df is None or len(df) == 0:
        return []
    return [{'timestamp': row['timestamp'], 'fuel': float(row['fuel']),
             'speed': float(row['speed']),
             'latitude': float(row.get('latitude', 0) or 0),
             'longitude': float(row.get('longitude', 0) or 0), 'adc': 0}
            for _, row in df.tail(n).iterrows()]

def get_car_state(car_id):
    with _car_states_meta_lock:
        state = car_states.get(car_id)
        if state is None:
            prefill = _baseline_points_for_car(car_id, WINDOW_SIZE)
            state = CarState(car_id, prefill_points=prefill)
            car_states[car_id] = state
            logger.info(f"🧩 [{car_id}] Buffer khởi tạo {len(prefill)} điểm nền" if prefill
                        else f"🧩 [{car_id}] Không có dữ liệu nền")
        return state


def process_window(window_df, car_id):
    df = window_df.copy()
    df['car_id'] = car_id
    df['segment_id'] = 1
    if 'timestamp' in df.columns:
        df['timestamp'] = pd.to_datetime(df['timestamp'])
    df = convert_gps_to_relative(df)
    df['final_label'] = 'Driving'
    feature_table = build_segment_features_table(df)
    feat_df = feature_table[FEATURE_COLS_21]
    feat_vec = feat_df.iloc[0].apply(pd.to_numeric, errors='coerce').fillna(0).values.astype(np.float32)
    seq_cols = ['fuel', 'speed', 'distance_step', 'bearing']
    seq = df[seq_cols].values.astype(np.float32)
    mask = np.ones(len(seq), dtype=bool)
    probs = model_wrapper.predict(seq, mask, feat_vec)
    fuel_vals = df['fuel'].values
    speed_vals = df['speed'].values
    duration_min = len(fuel_vals) * 5
    features = {
        'MovingAvg': float(np.mean(fuel_vals)),
        'RollingStd': float(np.std(fuel_vals)),
        'RegressionSlope': float(feat_vec[7]),
        'FuelRate': float((fuel_vals[-1] - fuel_vals[0]) / max(duration_min * 60, 1)),
        'WindowFuelDiff': float(feat_vec[4]),
        'MaxJump': float(max(0, feat_vec[10], feat_vec[11])),
        'StopDuration': int(duration_min * feat_vec[-1] * 60),
        'AvgSpeed': float(np.mean(speed_vals)),
        'FuelRange': float(feat_vec[14]),
    }
    return probs, features


def _normalize_incoming_point(p):
    try:
        ts = pd.to_datetime(p['timestamp'])
    except Exception:
        ts = pd.Timestamp(datetime.now())
    return {
        'timestamp': ts,
        'fuel_raw': float(p['fuel']),
        'fuel': float(p['fuel']),
        'speed': float(p.get('speed', 0)),
        'latitude': float(p.get('latitude', 0)),
        'longitude': float(p.get('longitude', 0)),
        'adc': p.get('ADC', p.get('adc', 0))
    }


window_manager = AdaptiveWindowManager()
model_wrapper = ModelWrapper(MODEL_PATH)


def _ingest_and_predict_one(point, car_id):
    state = get_car_state(car_id)
    with state.lock:
        state.data_buffer_raw.append({
            'timestamp': point['timestamp'],
            'fuel': point['fuel_raw'],
            'speed': point['speed'],
            'latitude': point['latitude'],
            'longitude': point['longitude'],
            'adc': point.get('adc', 0)
        })

        fuel_filtered = state.noise_filter.filter_fuel(point['fuel_raw'], point['speed'])
        retroactive_fuel = state.noise_filter._retroactive_point
        if retroactive_fuel is not None and state.data_buffer_filtered:
            previous_filtered = state.data_buffer_filtered[-1]
            previous_filtered['fuel'] = float(retroactive_fuel)
            if state.processed_buffer:
                state.processed_buffer[-1]['Fuel'] = float(retroactive_fuel)
        speed_filtered = state.noise_filter.filter_speed(point['speed'])
        delta_magnitude, signal_type = state.delta_detector.update(point['fuel_raw'], point['speed'], filtered_fuel=fuel_filtered)
        if delta_magnitude is not None:
            logger.info(f"🔎 [{car_id}] DeltaDetector: Δ={delta_magnitude:.2f}L, signal={signal_type}, raw={point['fuel_raw']:.2f}, filt={fuel_filtered:.2f}")
        else:
            logger.info(f"🔎 [{car_id}] DeltaDetector: Δ=NA, signal={signal_type}, raw={point['fuel_raw']:.2f}, filt={fuel_filtered:.2f}")

        # [FIX V3] Chặn 6 FP khi Driving nhanh: spike lớn nhưng baseline không lệch đáng kể
        if (signal_type in ('refuel', 'theft')
                and point['speed'] > 15.0
                and state.tracker.status == 'normal'
                and len(state.data_buffer_raw) >= 5):
            baseline = float(np.median([p['fuel'] for p in list(state.data_buffer_raw)[-5:-1]]))
            if abs(point['fuel_raw'] - baseline) < 25.0:
                signal_type = 'driving_noise'
                logger.info(f"⛔ [{car_id}] FP chặn: speed={point['speed']:.1f} raw={point['fuel_raw']:.1f} baseline={baseline:.1f}")

        window_buffer = list(state.data_buffer_filtered)
        window_buffer.append({
            'timestamp': point['timestamp'],
            'fuel': fuel_filtered,
            'speed': speed_filtered,
        })

        if signal_type in ('refuel', 'theft'):
            desired_window = max(4, window_manager.get_window_size(state.tracker, window_buffer))
        elif signal_type == 'pending':
            desired_window = max(4, window_manager.get_window_size(state.tracker, window_buffer))
        else:
            desired_window = window_manager.get_window_size(state.tracker, window_buffer)
            if state.tracker.status == 'normal' and len(window_buffer) >= 3:
                recent = window_buffer[-3:]
                fuels = [p['fuel'] for p in recent]
                if len(fuels) >= 3 and abs(fuels[-1] - fuels[-2]) > 10.0 and abs(fuels[-2] - fuels[-3]) < 2.0:
                    desired_window = max(6, desired_window)

        state.tracker._current_window = desired_window
        state.noise_filter.current_window = desired_window

        if signal_type in ('refuel', 'theft'):
            logger.info(f"⚡ [{car_id}] Confirmed {signal_type} → persisted signal; keeping filtered fuel (no raw passthrough)")

        filtered_point = {
            'timestamp': point['timestamp'],
            'fuel': fuel_filtered,
            'speed': speed_filtered,
            'latitude': point['latitude'],
            'longitude': point['longitude'],
            'adc': point.get('adc', 0),
            'fuel_raw': point['fuel_raw'],
            'speed_raw': point['speed']
        }
        state.data_buffer_filtered.append(filtered_point)

        logger.info(f"🔍 [{car_id}] Filter: RAW={point['fuel_raw']:.2f}L → FILTERED={fuel_filtered:.2f}L "
                    f"| Δ={fuel_filtered - point['fuel_raw']:+.2f}L | W={desired_window}")

        post_refuel_override = False
        if getattr(state.tracker, '_post_refuel_ticks', 0) > 0:
            state.tracker._post_refuel_ticks -= 1
            curr_speed = point['speed']
            fuel_delta_check = 0.0
            if len(state.data_buffer_filtered) >= 2:
                buf = list(state.data_buffer_filtered)
                fuel_delta_check = abs(buf[-1]['fuel'] - buf[-2]['fuel'])
            
            # IMPROVEMENT #3: Smart post-refuel skip - check fuel stability
            state.tracker._fuel_deltas_recent.append(fuel_delta_check)
            fuel_stability = 0.0
            if len(state.tracker._fuel_deltas_recent) >= 2:
                fuel_stability = max(0.0, 1.0 - (sum(state.tracker._fuel_deltas_recent) / len(state.tracker._fuel_deltas_recent)))
            
            # Early exit if fuel stabilizes: std(recent_deltas) < 0.5 and speed < 1.0
            if curr_speed < 1.0 and fuel_delta_check < 3.0 and fuel_stability > 0.85:
                # Fuel is stable: re-enable Idle/Theft detection early
                post_refuel_override = True
                logger.info(f"🔄 [{car_id}] post_refuel_ticks={state.tracker._post_refuel_ticks + 1} → fuel_stability={fuel_stability:.2f} → override Idle early")
            elif curr_speed < 1.0 and fuel_delta_check < 3.0:
                post_refuel_override = True
                logger.info(f"🔄 [{car_id}] post_refuel_ticks={state.tracker._post_refuel_ticks + 1} → override Idle")
            
            if state.tracker._post_refuel_ticks == 0:
                state.tracker._post_refuel_flag = False
                state.tracker._fuel_deltas_recent.clear()  # Reset for next refuel

        buffer_size = len(state.data_buffer_filtered)
        window_df, window_size = SegmentBuilder.build(state.data_buffer_filtered, desired_window)

        logger.info(f"📥 [{car_id}] | time={_ts_to_iso(point['timestamp'])} "
                    f"| fuel_raw={point['fuel_raw']:.2f} | fuel_filt={fuel_filtered:.2f} "
                    f"| buffer={buffer_size} | W={desired_window} | tracker={state.tracker.status}")

        probs, feats = process_window(window_df, car_id)

        if post_refuel_override:
            probs = np.zeros_like(probs)
            probs[LABELS.index('Idle')] = 1.0

        # A persisted fuel transition is stronger than a stale model window.
        # The tracker will backtrack to the pending point after confirmation.
        if signal_type in ('refuel', 'theft'):
            probs = np.full_like(probs, 0.02)
            probs[LABELS.index('Refuel' if signal_type == 'refuel' else 'Theft')] = 0.94

        current_speed = float(point['speed'])
        if current_speed <= 1.0:
            state.stationary_ticks += 1
            state.moving_ticks = 0
        elif current_speed > DRIVING_SPEED_GATE:
            state.moving_ticks += 1
            state.stationary_ticks = 0
        else:
            state.stationary_ticks = 0
            state.moving_ticks = 0

        active_event = state.tracker.event_record.get('event') if state.tracker.event_record else None
        protected_event = active_event in ('Refuel', 'Theft')
        raw_fuel_delta = 0.0
        if len(state.data_buffer_raw) >= 2:
            raw_points = list(state.data_buffer_raw)
            raw_fuel_delta = float(raw_points[-1]['fuel'] - raw_points[-2]['fuel'])
        fuel_transition = (
            current_speed <= 5.0
            and (raw_fuel_delta >= AdaptiveNoiseFilter.REFUEL_START_JUMP
                 or raw_fuel_delta <= -THEFT_MIN_DELTA)
        )
        if (not post_refuel_override and state.stationary_ticks >= 2
                and not protected_event and not fuel_transition):
            probs = np.zeros_like(probs)
            probs[LABELS.index('Idle')] = 1.0
            logger.info(f"🛑 [{car_id}] Motion gate: {state.stationary_ticks} stationary ticks → Idle")
        elif state.moving_ticks >= 1 and (protected_event or active_event == 'Idle'):
            probs = np.zeros_like(probs)
            probs[LABELS.index('Driving')] = 1.0
            logger.info(f"🚗 [{car_id}] Motion gate: speed={current_speed:.1f} → Driving")

        TIE_MARGIN = 0.15
        if not post_refuel_override and len(state.data_buffer_filtered) >= 2:
            recent2 = list(state.data_buffer_filtered)[-2:]
            delta_fuel = recent2[-1]['fuel'] - recent2[0]['fuel']
            sorted_idx = np.argsort(probs)[::-1]
            margin = float(probs[sorted_idx[0]] - probs[sorted_idx[1]])
            if margin < TIE_MARGIN:
                if delta_fuel <= -THEFT_MIN_DELTA and 'Theft' in (LABELS[sorted_idx[0]], LABELS[sorted_idx[1]]):
                    probs = np.full_like(probs, 0.02)
                    probs[LABELS.index('Theft')] = 1.0 - 0.02 * (len(LABELS) - 1)
                elif delta_fuel >= REFUEL_MIN_DELTA and 'Refuel' in (LABELS[sorted_idx[0]], LABELS[sorted_idx[1]]):
                    probs = np.full_like(probs, 0.02)
                    probs[LABELS.index('Refuel')] = 1.0 - 0.02 * (len(LABELS) - 1)

        if not post_refuel_override and len(state.data_buffer_filtered) >= 2:
            recent = list(state.data_buffer_filtered)[-2:]
            speeds = [p['speed'] for p in recent]
            fuels = [p['fuel'] for p in recent]
            if max(speeds) < 1.0 and (max(fuels) - min(fuels)) < 0.3:
                probs = np.zeros_like(probs)
                probs[LABELS.index('Idle')] = 1.0

        TRANSITION_HOLD_CONF = 0.65
        TRANSITION_HOLD_MARGIN = 0.15
        if (not post_refuel_override
                and state.tracker.status == 'confirmed'
                and state.tracker.event_record
                and state.tracker.event_record.get('event') not in ('Refuel', 'Theft')):
            sorted_idx = np.argsort(probs)[::-1]
            top_conf = float(probs[sorted_idx[0]])
            margin = float(probs[sorted_idx[0]] - probs[sorted_idx[1]])
            runner_up_label = LABELS[sorted_idx[1]]
            skip_hold = False
            if len(state.data_buffer_filtered) >= 2:
                recent2 = list(state.data_buffer_filtered)[-2:]
                delta_fuel = recent2[-1]['fuel'] - recent2[0]['fuel']
                if runner_up_label == 'Theft' and delta_fuel <= -THEFT_MIN_DELTA:
                    skip_hold = True
                elif runner_up_label == 'Refuel' and delta_fuel >= REFUEL_MIN_DELTA:
                    skip_hold = True
            if not skip_hold and (top_conf < TRANSITION_HOLD_CONF or margin < TRANSITION_HOLD_MARGIN):
                held_label = state.tracker.event_record['event']
                probs = np.full_like(probs, 0.02)
                probs[LABELS.index(held_label)] = 1.0 - 0.02 * (len(LABELS) - 1)

        pred_label = LABELS[int(np.argmax(probs))]
        conf = float(probs[int(np.argmax(probs))])
        probs_str = ", ".join([f"{LABELS[i]}={probs[i]:.3f}" for i in range(4)])
        logger.info(f"🤖 [{car_id}] Dự đoán: {pred_label} (conf={conf:.3f}) | W={window_size} | [{probs_str}]")

        full_df = pd.DataFrame(list(state.data_buffer_filtered))
        full_df['timestamp'] = pd.to_datetime(full_df['timestamp'])
        state.tracker.update(probs, full_df, point['timestamp'])

        tracker_status = state.tracker.status
        event_type = None
        if tracker_status in ('candidate', 'confirmed'):
            event_type = state.tracker.candidate_label
            if not event_type and state.tracker.event_record:
                event_type = state.tracker.event_record.get('event')
        if getattr(state.tracker, '_pending_switch', None):
            event_type = state.tracker._pending_switch.get('label')
        state.noise_filter.set_tracker_status(tracker_status, event_type)

        display_stage = state.tracker.display_stage
        point_status = state.tracker.point_status
        logger.info(f"📊 [{car_id}] Tracker: {tracker_status} | {display_stage} | {point_status}")

        finished_event = state.tracker.pop_finished_event()
        if finished_event:
            state.event_history.append(finished_event)
            _append_event_log(finished_event, car_id)
            logger.info(f"🎯 [{car_id}] EVENT FINISHED: {finished_event.get('event')} | "
                        f"start={finished_event.get('start_time')} | end={finished_event.get('end_time')} | "
                        f"fuel_change={finished_event.get('fuel_change')}L | duration={finished_event.get('duration_s')}s")

            state.noise_filter.reset(soft=True)
            anchor_val = None
            if state.noise_filter.filtered_history:
                anchor_val = float(state.noise_filter.filtered_history[-1])
                state.delta_detector.anchor_ema(anchor_val)
            state.delta_detector._reset()
            if anchor_val is not None:
                logger.info(f"🔄 [{car_id}] Soft reset filter + EMA anchor={anchor_val:.2f}L sau event")
            else:
                logger.info(f"🔄 [{car_id}] Soft reset filter + EMA reset sau event")

        if tracker_status == 'candidate':
            logger.info(f"  ⏳ [{car_id}] Candidate: {state.tracker.candidate_label} "
                        f"(counter={state.tracker.counter}/{CONFIRM_COUNT}, "
                        f"mismatch={state.tracker.mismatch_streak}, "
                        f"thresh={state.tracker._get_confirm_threshold():.3f})")

        if tracker_status == 'confirmed' and state.tracker._just_confirmed:
            logger.info(f"  ✅ [{car_id}] CONFIRMED: {state.tracker.event_record.get('event')} "
                        f"| confidence={state.tracker.best_conf:.3f}")
            state.tracker.retroactively_fix_point_status(full_df, state.processed_buffer)

        last = window_df.iloc[-1]
        record = {
            'Timestamp': point['timestamp'],
            'Fuel': last['fuel'],
            'Fuel_Raw': point['fuel_raw'],
            'Fuel_Delta': fuel_filtered - point['fuel_raw'],
            'Speed': last['speed'],
            'ADC': last.get('adc', 0),
            'Latitude': last['latitude'],
            'Longitude': last['longitude'],
            'Prediction': pred_label,
            'PointStatus': point_status,
            'Confidence': conf * 100,
            'WindowSize': window_size,
            'BoundaryPoint': bool(state.tracker.event_record.get('boundary_point')) if state.tracker.event_record else False,
            **feats
        }
        state.processed_buffer.append(record)
        if len(state.processed_buffer) > MAX_POINTS_HISTORY:
            del state.processed_buffer[0: len(state.processed_buffer) - MAX_POINTS_HISTORY]
        _save_point_to_db(car_id, record)

        return make_json_safe({
            'car_id': car_id,
            'status': point_status,
            'record': {
                'timestamp': record['Timestamp'].isoformat(),
                'fuel': record['Fuel'],
                'fuel_raw': record['Fuel_Raw'],
                'fuel_delta': record['Fuel_Delta'],
                'speed': record['Speed'],
                'prediction': record['Prediction'],
                'point_status': point_status,
                'confidence': record['Confidence'],
                'window_size': window_size
            },
            'tracker_status': tracker_status,
            'active_event': state.tracker.event_record
        })


def visible_cars():
    user = session.get("user", {})
    if user.get("role") == "admin":
        return sorted(set(car_data_cache) | set(car_states))
    car_id = user.get("car_id")
    return [car_id] if car_id and car_id in car_data_cache else []

def authorize_car(car_id):
    user = session.get("user", {})
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if car_id is None and user.get("role") == "admin":
        return None
    if car_id == "All" and user.get("role") == "admin":
        return None
    if car_id not in car_data_cache and car_id not in car_states:
        return jsonify({"error": "Car not found"}), 404
    if user.get("role") != "admin" and user.get("car_id") != car_id:
        return jsonify({"error": "forbidden"}), 403
    return None


@app.route('/')
def home():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        remember = request.form.get("remember") == "on"
        if not email or not password:
            flash("Vui lòng nhập email và mật khẩu!", "danger")
            return redirect(url_for("login"))
        try:
            result = supabase.table("accounts").select("id, email, password").eq("email", email).limit(1).execute()
            if not result.data:
                flash("Email hoặc mật khẩu không chính xác!", "danger")
                return redirect(url_for("login"))
            account = result.data[0]
            if not check_password_hash(account["password"], password):
                flash("Email hoặc mật khẩu không chính xác!", "danger")
                return redirect(url_for("login"))
            user_info_result = supabase.table("user_info").select("full_name, role, car_id").eq(
                "account_id", account["id"]).limit(1).execute()
            full_name = user_info_result.data[0].get("full_name", "") if user_info_result.data else ""
            role = user_info_result.data[0].get("role", "viewer") if user_info_result.data else "viewer"
            car_id = user_info_result.data[0].get("car_id") if user_info_result.data else None
            session.clear()
            session["user"] = {"id": account["id"], "email": account["email"],
                               "full_name": full_name, "role": role, "car_id": car_id}
            session.permanent = bool(remember)
            flash("Đăng nhập thành công!", "success")
            return redirect(url_for("dashboard"))
        except Exception as e:
            logger.error(f"Login error: {e}")
            flash("Có lỗi xảy ra khi đăng nhập!", "danger")
            return redirect(url_for("login"))
    return render_template("login.html")

@app.route("/dashboard")
@login_required
def dashboard():
    user = session["user"]
    template = "dashboard.html" if user.get("role") == "admin" else "dashboard_user.html"
    return render_template(template, user=user, cars=visible_cars(),
                           current_car=user.get("car_id") or globals()["current_car"])

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        if not full_name:
            flash("Vui lòng nhập họ và tên!", "danger")
            return redirect(url_for("register"))
        if not email:
            flash("Vui lòng nhập email!", "danger")
            return redirect(url_for("register"))
        if len(password) < 8:
            flash("Mật khẩu phải có ít nhất 8 ký tự!", "danger")
            return redirect(url_for("register"))
        if password != confirm_password:
            flash("Mật khẩu xác nhận không khớp!", "danger")
            return redirect(url_for("register"))
        try:
            existing = supabase.table("accounts").select("id").eq("email", email).limit(1).execute()
            if existing.data:
                flash("Email này đã được đăng ký!", "danger")
                return redirect(url_for("register"))
            password_hash = generate_password_hash(password)
            account_result = supabase.table("accounts").insert({"email": email, "password": password_hash}).execute()
            if not account_result.data:
                flash("Không thể tạo tài khoản!", "danger")
                return redirect(url_for("register"))
            account_id = account_result.data[0]["id"]
            info_result = supabase.table("user_info").insert({"account_id": account_id, "full_name": full_name}).execute()
            if not info_result.data:
                supabase.table("accounts").delete().eq("id", account_id).execute()
                flash("Không thể lưu thông tin người dùng!", "danger")
                return redirect(url_for("register"))
            flash("Đăng ký tài khoản thành công! Mời bạn đăng nhập.", "success")
            return redirect(url_for("login"))
        except Exception as e:
            logger.error(f"Register error: {e}")
            flash(f"Lỗi đăng ký: {str(e)}", "danger")
            return redirect(url_for("register"))
    return render_template("register.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Bạn đã đăng xuất khỏi FuelSentinel AI.", "success")
    return redirect(url_for("login"))

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if not email:
            flash("Vui lòng nhập email!", "danger")
            return redirect(url_for("forgot_password"))
        try:
            result = supabase.table("accounts").select("id, email").eq("email", email).limit(1).execute()
            if not result.data:
                flash("Email không tồn tại trong hệ thống!", "danger")
                return redirect(url_for("forgot_password"))
            otp = str(random.randint(100000, 999999))
            session["reset_email"] = email
            session["reset_otp"] = otp
            session["reset_otp_expire"] = (datetime.now() + timedelta(minutes=5)).timestamp()
            session["otp_verified"] = False
            send_otp_email(email, otp)
            flash("Mã OTP đã được gửi đến email của bạn!", "success")
            return redirect(url_for("verify_otp"))
        except Exception as e:
            logger.error(f"Forgot password error: {e}")
            flash("Không thể gửi mã OTP. Vui lòng thử lại!", "danger")
            return redirect(url_for("forgot_password"))
    return render_template("forgot_password.html")

@app.route("/verify-otp", methods=["GET", "POST"])
def verify_otp():
    if "reset_email" not in session:
        flash("Phiên đặt lại mật khẩu không hợp lệ!", "danger")
        return redirect(url_for("forgot_password"))
    if request.method == "POST":
        otp_input = request.form.get("otp", "").strip()
        saved_otp = session.get("reset_otp")
        expire_time = session.get("reset_otp_expire")
        if not saved_otp or not expire_time:
            flash("Mã OTP không hợp lệ!", "danger")
            return redirect(url_for("forgot_password"))
        if datetime.now().timestamp() > expire_time:
            session.pop("reset_otp", None)
            session.pop("reset_otp_expire", None)
            flash("Mã OTP đã hết hạn!", "danger")
            return redirect(url_for("forgot_password"))
        if otp_input != saved_otp:
            flash("Mã OTP không chính xác!", "danger")
            return redirect(url_for("verify_otp"))
        session["otp_verified"] = True
        session.pop("reset_otp", None)
        session.pop("reset_otp_expire", None)
        flash("Xác minh OTP thành công!", "success")
        return redirect(url_for("reset_password"))
    return render_template("verify_otp.html")

@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    if not session.get("otp_verified"):
        flash("Vui lòng xác minh OTP trước!", "danger")
        return redirect(url_for("forgot_password"))
    email = session.get("reset_email")
    if not email:
        flash("Phiên đặt lại mật khẩu không hợp lệ!", "danger")
        return redirect(url_for("forgot_password"))
    if request.method == "POST":
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        if len(password) < 8:
            flash("Mật khẩu phải có ít nhất 8 ký tự!", "danger")
            return redirect(url_for("reset_password"))
        if password != confirm_password:
            flash("Mật khẩu xác nhận không khớp!", "danger")
            return redirect(url_for("reset_password"))
        try:
            password_hash = generate_password_hash(password)
            result = supabase.table("accounts").update({"password": password_hash}).eq("email", email).execute()
            if not result.data:
                flash("Không thể cập nhật mật khẩu!", "danger")
                return redirect(url_for("reset_password"))
            session.pop("reset_email", None)
            session.pop("otp_verified", None)
            flash("Đặt lại mật khẩu thành công! Bạn có thể đăng nhập bằng mật khẩu mới.", "success")
            return redirect(url_for("login"))
        except Exception as e:
            logger.error(f"Reset password error: {e}")
            flash("Có lỗi xảy ra khi đổi mật khẩu!", "danger")
            return redirect(url_for("reset_password"))
    return render_template("reset_password.html")

def send_otp_email(receiver_email, otp):
    try:
        message = EmailMessage()
        message["Subject"] = "FuelSentinel - Ma xac minh dat lai mat khau"
        message["From"] = GMAIL_EMAIL
        message["To"] = receiver_email
        message.set_content(f"Ma OTP cua ban la: {otp}\nCo hieu luc trong 5 phut.")
        context = ssl.create_default_context(cafile=certifi.where())
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls(context=context)
            server.login(GMAIL_EMAIL, GMAIL_APP_PASSWORD)
            server.send_message(message)
    except Exception as e:
        logger.error(f"Send OTP error: {e}")
        raise

@app.route('/dataset')
@login_required
def dataset():
    cars = visible_cars()
    return render_template('dataset.html', cars=cars, current_car=cars[0] if len(cars) == 1 else current_car)

@app.route('/feature')
@login_required
def feature():
    cars = visible_cars()
    return render_template('feature.html', cars=cars, current_car=cars[0] if len(cars) == 1 else current_car)

@app.route('/realtime')
@login_required
def realtime():
    cars = visible_cars()
    return render_template('realtime.html', cars=cars, current_car=cars[0] if len(cars) == 1 else current_car)

@app.route('/events')
@login_required
def events():
    cars = visible_cars()
    return render_template('events.html', cars=cars, current_car=cars[0] if len(cars) == 1 else current_car)

@app.route('/report')
@login_required
def report_page():
    cars = visible_cars()
    return render_template('report.html', cars=cars, current_car=cars[0] if len(cars) == 1 else current_car)

@app.route('/settings')
@login_required
def settings():
    cars = visible_cars()
    return render_template('settings.html', cars=cars, current_car=cars[0] if len(cars) == 1 else current_car)

@app.route('/change-password')
@login_required
def change_password_page():
    cars = visible_cars()
    return render_template('change_password.html', cars=cars, current_car=cars[0] if len(cars) == 1 else None)

@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    try:
        data = request.get_json(silent=True) or {}
        current_password = str(data.get('current_password', '')).strip()
        new_password = str(data.get('new_password', ''))
        confirm_password = str(data.get('confirm_password', ''))
        if not current_password:
            return jsonify({'status': 'error', 'message': 'Vui lòng nhập mật khẩu hiện tại.'}), 400
        if not new_password:
            return jsonify({'status': 'error', 'message': 'Vui lòng nhập mật khẩu mới.'}), 400
        if not confirm_password:
            return jsonify({'status': 'error', 'message': 'Vui lòng xác nhận mật khẩu mới.'}), 400
        if len(new_password) < 8:
            return jsonify({'status': 'error', 'message': 'Mật khẩu phải có ít nhất 8 ký tự.'}), 400
        if new_password != confirm_password:
            return jsonify({'status': 'error', 'message': 'Mật khẩu xác nhận không khớp.'}), 400
        user = session.get('user', {})
        email = user.get('email')
        if not email:
            return jsonify({'status': 'error', 'message': 'Không xác định được tài khoản.'}), 400
        result = supabase.table('accounts').select('password').eq('email', email).limit(1).execute()
        if not result.data:
            return jsonify({'status': 'error', 'message': 'Không tìm thấy tài khoản.'}), 404
        password_hash = result.data[0].get('password')
        if not password_hash:
            return jsonify({'status': 'error', 'message': 'Tài khoản chưa có mật khẩu hợp lệ.'}), 400
        if not check_password_hash(password_hash, current_password):
            return jsonify({'status': 'error', 'message': 'Mật khẩu hiện tại không chính xác.'}), 400
        if check_password_hash(password_hash, new_password):
            return jsonify({'status': 'error', 'message': 'Mật khẩu mới không được trùng mật khẩu hiện tại.'}), 400
        new_password_hash = generate_password_hash(new_password)
        update_result = supabase.table('accounts').update({'password': new_password_hash}).eq('email', email).execute()
        if not update_result.data:
            return jsonify({'status': 'error', 'message': 'Không thể cập nhật mật khẩu.'}), 500
        return jsonify({'status': 'success', 'message': 'Đổi mật khẩu thành công.'}), 200
    except Exception as e:
        logger.error(f"Change password error: {e}")
        return jsonify({'status': 'error', 'message': 'Có lỗi xảy ra khi đổi mật khẩu.'}), 500

@app.route('/api/set_vehicle', methods=['POST'])
@login_required
def set_vehicle():
    data = request.json or {}
    car_id = data.get('car_id')
    if not car_id:
        return jsonify({'error': 'Missing car_id'}), 400
    error = authorize_car(car_id)
    if error:
        return error
    session['selected_car'] = car_id
    get_car_state(car_id)
    return jsonify({'status': 'ok', 'car_id': car_id})

@app.route('/api/current_vehicle')
@login_required
def current_vehicle():
    user = session['user']
    return jsonify({'car_id': session.get('selected_car') if user.get('role') == 'admin' else user.get('car_id')})

@app.route('/api/car_data')
@login_required
def api_car_data():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    limit = int(request.args.get('limit', 500))
    error = authorize_car(car_id)
    if error:
        return error
    if car_id not in car_data_cache:
        return jsonify({'error': 'Car not found'}), 404
    df = car_data_cache[car_id].tail(limit).copy()
    data = []
    for _, row in df.iterrows():
        data.append(make_json_safe({
            'timestamp': row['timestamp'].isoformat(),
            'fuel': float(row['fuel']),
            'fuel_filter': float(row.get('fuel_filter', row['fuel'])),
            'speed': float(row['speed']),
            'latitude': float(row['latitude']),
            'longitude': float(row['longitude']),
            'label': row.get('final_label', 'Driving')
        }))
    return jsonify(data)

@app.route('/api/predict', methods=['POST'])
def api_predict():
    payload = request.get_json(silent=True)
    if payload is None:
        return jsonify({'error': 'Missing/invalid JSON body'}), 400
    raw_points = payload if isinstance(payload, list) else [payload]
    if len(raw_points) == 0:
        return jsonify({'error': 'Empty payload'}), 400
    results = []
    for p in raw_points:
        car_id = p.get('car_id') or p.get('carId')
        if not car_id:
            return jsonify({'error': "Missing 'car_id' in point payload"}), 400
        try:
            point = _normalize_incoming_point(p)
        except (KeyError, TypeError, ValueError) as e:
            return jsonify({'error': f'Invalid point payload: {e}'}), 400
        try:
            res = _ingest_and_predict_one(point, car_id)
        except Exception as e:
            traceback.print_exc()
            return jsonify({'error': f'Predict failed: {e}'}), 500
        results.append(res)
    return jsonify(results if isinstance(payload, list) else results[0])

@app.route('/api/raw_points')
@login_required
def api_raw_points():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    limit = int(request.args.get('limit', 50))
    state = get_car_state(car_id)
    with state.lock:
        raw_buffer = list(state.data_buffer_raw)
        filtered_buffer = list(state.data_buffer_filtered)
        processed_buffer = list(state.processed_buffer)
    raw_recent = raw_buffer[-limit:]
    filtered_recent = filtered_buffer[-limit:]
    points = []
    for i in range(max(len(raw_recent), len(filtered_recent))):
        raw = raw_recent[i] if i < len(raw_recent) else None
        filt = filtered_recent[i] if i < len(filtered_recent) else None
        proc = processed_buffer[i] if i < len(processed_buffer) else None
        points.append({
            'timestamp': raw['timestamp'].isoformat() if raw else None,
            'raw': {'fuel': raw['fuel'] if raw else None, 'speed': raw['speed'] if raw else None},
            'filtered': {'fuel': filt['fuel'] if filt else None, 'speed': filt['speed'] if filt else None},
            'processed': {
                'prediction': proc['Prediction'] if proc else None,
                'point_status': proc['PointStatus'] if proc else None,
                'confidence': proc['Confidence'] if proc else None,
            }
        })
    return jsonify({'car_id': car_id, 'count': len(points), 'points': points})

@app.route('/api/points')
@login_required
def api_points():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    limit = max(1, min(int(request.args.get('limit', DEFAULT_VISIBLE_POINTS)), MAX_POINTS_HISTORY))
    offset = max(0, int(request.args.get('offset', 0)))
    state = get_car_state(car_id)
    with state.lock:
        buf = list(state.processed_buffer)
    total = len(buf)
    end = max(0, total - offset)
    start = max(0, end - limit)
    page = buf[start:end]
    points = []
    for item in page:
        points.append(make_json_safe({
            'timestamp': item['Timestamp'].isoformat() if hasattr(item['Timestamp'], 'isoformat') else str(item['Timestamp']),
            'fuel': item['Fuel'],
            'fuel_raw': item.get('Fuel_Raw'),
            'fuel_delta': item.get('Fuel_Delta'),
            'speed': item['Speed'],
            'prediction': item['Prediction'],
            'point_status': item.get('PointStatus', 'normal'),
            'confidence': item['Confidence'],
            'window_size': item.get('WindowSize'),
            'boundary_point': item.get('BoundaryPoint', False)
        }))
    return jsonify({'car_id': car_id, 'total': total, 'limit': limit, 'offset': offset, 'points': points})

@app.route('/api/history')
@login_required
def api_history():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    with state.lock:
        history = state.processed_buffer[-100:] if state.processed_buffer else []
    result = []
    for item in history:
        result.append(make_json_safe({
            "raw": {
                "Timestamp": item['Timestamp'].isoformat() if hasattr(item['Timestamp'], 'isoformat') else str(item['Timestamp']),
                "Fuel": item['Fuel'], "Fuel_Raw": item.get('Fuel_Raw'),
                "Fuel_Delta": item.get('Fuel_Delta'), "Speed": item['Speed'],
                "ADC": item.get('ADC', 0), "Latitude": item['Latitude'], "Longitude": item['Longitude']
            },
            "processed": {
                "Prediction": item['Prediction'], "PointStatus": item.get('PointStatus', 'normal'),
                "RegressionSlope": item['RegressionSlope'], "RegressionR2": 0.0,
                "FuelRate": item['FuelRate'], "InstantFuelDiff": item.get('WindowFuelDiff', 0),
                "WindowFuelDiff": item.get('WindowFuelDiff', 0), "MaxJump": item.get('MaxJump', 0),
                "MovingAvg": item['MovingAvg'], "RollingStd": item['RollingStd'],
                "StopDuration": item['StopDuration'], "AvgSpeed": item['AvgSpeed'],
                "FuelRange": item['FuelRange'], "Confidence": item['Confidence'],
                "WindowSize": item.get('WindowSize')
            }
        }))
    return jsonify(result)

@app.route('/api/realtime')
@login_required
def api_realtime():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    with state.lock:
        if not state.processed_buffer:
            return jsonify(make_json_safe({
                "raw": {"Timestamp": datetime.now().isoformat(), "Fuel": 0, "Fuel_Raw": 0,
                        "Speed": 0, "ADC": 0, "Latitude": 0, "Longitude": 0},
                "processed": {"Prediction": "Idle", "PointStatus": "normal",
                              "RegressionSlope": 0, "FuelRate": 0, "MovingAvg": 0,
                              "RollingStd": 0, "StopDuration": 0, "AvgSpeed": 0,
                              "FuelRange": 0, "Confidence": 0, "WindowSize": 0}
            }))
        last = state.processed_buffer[-1]
    return jsonify(make_json_safe({
        "raw": {
            "Timestamp": last['Timestamp'].isoformat() if hasattr(last['Timestamp'], 'isoformat') else str(last['Timestamp']),
            "Fuel": last['Fuel'], "Fuel_Raw": last.get('Fuel_Raw'), "Fuel_Delta": last.get('Fuel_Delta'),
            "Speed": last['Speed'], "ADC": last.get('ADC', 0),
            "Latitude": last['Latitude'], "Longitude": last['Longitude']
        },
        "processed": {
            "Prediction": last['Prediction'], "PointStatus": last.get('PointStatus', 'normal'),
            "RegressionSlope": last['RegressionSlope'], "FuelRate": last['FuelRate'],
            "MovingAvg": last['MovingAvg'], "RollingStd": last['RollingStd'],
            "StopDuration": last['StopDuration'], "AvgSpeed": last['AvgSpeed'],
            "FuelRange": last['FuelRange'], "Confidence": last['Confidence'],
            "WindowSize": last.get('WindowSize')
        }
    }))

@app.route('/api/events')
@login_required
def api_events():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    with state.lock:
        evts = list(state.event_history)
    result = []
    for evt in evts:
        result.append(make_json_safe({
            'state': evt.get('event', 'Unknown'),
            'start_time': evt['start_time'].isoformat() if hasattr(evt['start_time'], 'isoformat') else str(evt['start_time']),
            'end_time': evt['end_time'].isoformat() if hasattr(evt['end_time'], 'isoformat') else str(evt['end_time']),
            'start_fuel': evt.get('fuel_before', 0), 'end_fuel': evt.get('fuel_after', 0),
            'delta': evt.get('fuel_change', 0), 'min_fuel': evt.get('min_fuel'),
            'fuel_added': evt.get('fuel_added', evt.get('fuel_change', 0)),
            'duration_s': evt.get('duration_s', 0), 'avg_rate': 0.0,
            'confidence': evt.get('confidence', 0) * 100,
            'max_confidence': evt.get('max_confidence', evt.get('confidence', 0)) * 100,
            'stage_history': evt.get('stage_history', [])
        }))
    return jsonify(result)

@app.route('/api/event_log')
@login_required
def api_event_log():
    limit = int(request.args.get('limit', 1000))
    car_id = request.args.get('car_id') or session['user'].get('car_id')
    error = authorize_car(car_id)
    if error:
        return error
    raw_logs = _read_event_log(limit=limit, car_id=car_id)
    result = []
    for evt in raw_logs:
        if not evt:
            continue
        duration_s = evt.get('duration_s', 0) or 0
        fuel_change = evt.get('fuel_change', evt.get('fuel_added', 0)) or 0
        avg_rate = round((fuel_change / duration_s) * 60, 4) if duration_s else 0.0
        raw_confidence = evt.get('confidence', 0) or 0
        raw_max_confidence = evt.get('max_confidence', raw_confidence) or 0
        result.append({
            'state': evt.get('event', 'Unknown'),
            'start_time': evt.get('start_time'), 'end_time': evt.get('end_time'),
            'start_fuel': evt.get('fuel_before', 0), 'end_fuel': evt.get('fuel_after', 0),
            'delta': fuel_change, 'min_fuel': evt.get('min_fuel'),
            'fuel_added': evt.get('fuel_added', fuel_change),
            'duration_s': duration_s, 'avg_rate': avg_rate,
            'confidence': raw_confidence * 100, 'max_confidence': raw_max_confidence * 100,
            'stage_history': evt.get('stage_history', [])
        })
    return jsonify({'total': len(result), 'events': result})

@app.route('/api/toasts')
@login_required
def api_toasts():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    return jsonify(_read_toast_log(limit=4, car_id=car_id))

@app.route('/api/alerts')
@login_required
def api_alerts():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    logs = _read_event_log(limit=50, car_id=car_id)
    active_event = None
    with state.lock:
        full_df = None
        if state.processed_buffer:
            full_df = pd.DataFrame(state.processed_buffer)
            full_df.rename(columns={'Timestamp': 'timestamp', 'Fuel': 'fuel', 'Speed': 'speed',
                                    'Prediction': 'prediction', 'PointStatus': 'point_status',
                                    'Confidence': 'confidence'}, inplace=True)
            full_df['timestamp'] = pd.to_datetime(full_df['timestamp'])
        if state.tracker.event_record:
            active_event = dict(state.tracker.event_record)
            if active_event.get('start_time'):
                active_event['start_time'] = _ts_to_iso(active_event['start_time'])
            if active_event.get('end_time'):
                active_event['end_time'] = _ts_to_iso(active_event['end_time'])
            active_summary = state.tracker.get_active_summary(full_df)
            if active_summary:
                active_event.update({
                    'fuel_current': active_summary.get('fuel_current'),
                    'fuel_start': active_summary.get('fuel_start'),
                    'delta_fuel': active_summary.get('delta_fuel'),
                    'fuel_status': active_summary.get('fuel_status'),
                    'duration_s': active_summary.get('duration_s')
                })
    return jsonify(make_json_safe({'history': logs, 'active': active_event}))

@app.route('/api/download_report', methods=['POST'])
@login_required
def api_download_report():
    data = request.get_json()
    car_id = data.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    event_type = data.get('event_type', 'All')
    raw_logs = _read_event_log(limit=50000, car_id=None if car_id == 'All' else car_id)
    filtered_events = []
    start_dt = pd.to_datetime(start_date + "T00:00:00") if start_date else None
    end_dt = pd.to_datetime(end_date + "T23:59:59") if end_date else None
    for ev in raw_logs:
        if not ev:
            continue
        ev_start = ev.get('start_time')
        if ev_start:
            try:
                ev_dt = pd.to_datetime(ev_start)
                if start_dt and ev_dt < start_dt:
                    continue
                if end_dt and ev_dt > end_dt:
                    continue
            except Exception:
                pass
        if event_type != 'All' and ev.get('event') != event_type:
            continue
        duration_s = ev.get('duration_s', 0) or 0
        fuel_change = ev.get('fuel_change', ev.get('fuel_added', 0)) or 0
        avg_rate = round((fuel_change / duration_s) * 60, 4) if duration_s else 0.0
        raw_max_confidence = ev.get('max_confidence', ev.get('confidence', 0)) or 0
        filtered_events.append({
            'state': ev.get('event', 'Unknown'),
            'start_time': ev.get('start_time'), 'end_time': ev.get('end_time'),
            'start_fuel': ev.get('fuel_before', 0), 'end_fuel': ev.get('fuel_after', 0),
            'delta': fuel_change, 'duration_s': duration_s,
            'avg_rate': avg_rate, 'max_confidence': raw_max_confidence * 100
        })

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo sự kiện"
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for merge_range, value, size in [
        ('A1:I1', "BÁO CÁO HÀNH TRÌNH XE CHẠY", 18),
        ('A2:I2', "(Theo QCVN 06:2024/BCA)", 12),
        ('A3:I3', f"Từ 0 giờ 0 phút ngày {start_date or '---'} đến 23 giờ 59 phút ngày {end_date or '---'}", 11),
        ('A4:I4', "Đơn vị kinh doanh vận tải: Demo sensor nhiên liệu", 11),
        ('A5:I5', f"Biển số: {car_id}", 11),
    ]:
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(':')[0]]
        cell.value = value
        cell.font = Font(bold=(size >= 12), size=size, name='Times New Roman')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_offset = 7
    headers = ["Thời gian bắt đầu", "Thời gian kết thúc", "Nhiên liệu bắt đầu (L)",
               "Nhiên liệu kết thúc (L)", "Thay đổi nhiên liệu (L)", "Thời lượng",
               "Tốc độ thay đổi (L/h)", "Loại sự kiện", "Độ tin cậy (%)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_offset, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, name='Times New Roman')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for idx, ev in enumerate(filtered_events, start=1):
        row = row_offset + idx
        s_str, e_str = ev.get('start_time', ''), ev.get('end_time', '')
        try:
            if s_str:
                s_str = pd.to_datetime(s_str).strftime('%H:%M:%S - %d/%m/%Y')
            if e_str:
                e_str = pd.to_datetime(e_str).strftime('%H:%M:%S - %d/%m/%Y')
        except Exception:
            pass
        dur_s = ev.get('duration_s', 0)
        dur_str = f"{int(dur_s // 3600)}h {int((dur_s % 3600) // 60)}m {int(dur_s % 60)}s" if dur_s else "--"
        values = [s_str, e_str, round(ev.get('start_fuel', 0), 2), round(ev.get('end_fuel', 0), 2),
                  round(ev.get('delta', 0), 2), dur_str,
                  round(ev.get('avg_rate', 0) * 60, 2), ev.get('state', ''),
                  round(ev.get('max_confidence', 0), 2)]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row, column=col_idx).value = val
            ws.cell(row=row, column=col_idx).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 18
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     download_name=f'FuelSentinel_Report_{car_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                     as_attachment=True)

@app.route('/api/dashboard')
@login_required
def api_dashboard():
    car_id = request.args.get("car_id", session["user"].get("car_id") or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    with state.lock:
        if not state.processed_buffer:
            return jsonify(make_json_safe({
                "vehicle": car_id, "fuel": 0, "fuel_raw": 0, "speed": 0,
                "status": "Idle", "point_status": "normal", "slope": 0,
                "r2": 0, "confidence": 0, "stop_duration": 0, "alerts": 0,
                "last_update": datetime.now().isoformat()
            }))
        last = state.processed_buffer[-1]
        alerts_count = len(state.event_history)
    return jsonify(make_json_safe({
        "vehicle": car_id,
        "fuel": last.get("Fuel", 0), "fuel_raw": last.get("Fuel_Raw", 0),
        "fuel_delta": last.get("Fuel_Delta", 0), "speed": last.get("Speed", 0),
        "status": last.get("Prediction", "Idle"), "point_status": last.get("PointStatus", "normal"),
        "slope": last.get("RegressionSlope", 0), "r2": last.get("R2", 0),
        "confidence": last.get("Confidence", 0), "stop_duration": last.get("StopDuration", 0),
        "alerts": alerts_count,
        "last_update": last["Timestamp"].isoformat() if hasattr(last.get("Timestamp"), "isoformat") else str(last.get("Timestamp", ""))
    }))

@app.route('/api/tracker_status')
@login_required
def api_tracker_status():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    tracker = state.tracker
    full_df = None
    with state.lock:
        if state.processed_buffer:
            full_df = pd.DataFrame(state.processed_buffer)
            full_df.rename(columns={'Timestamp': 'timestamp', 'Fuel': 'fuel', 'Speed': 'speed',
                                    'Prediction': 'prediction', 'PointStatus': 'point_status',
                                    'Confidence': 'confidence'}, inplace=True)
            full_df['timestamp'] = pd.to_datetime(full_df['timestamp'])
        current_window = window_manager.get_window_size(tracker)
    active_summary = tracker.get_active_summary(full_df)
    return jsonify(make_json_safe({
        'car_id': car_id,
        'status': tracker.status,
        'point_status': tracker.point_status,
        'display_stage': tracker.display_stage,
        'current_label': (tracker.candidate_label if tracker.status == 'candidate'
                          else (tracker.event_record['event'] if tracker.event_record
                                else (tracker.last_prediction or 'Normal'))),
        'last_prediction': tracker.last_prediction,
        'confidence': tracker.best_conf,
        'current_window_size': current_window,
        'confirm_threshold': tracker._get_confirm_threshold(),
        'post_refuel_ticks': getattr(tracker, '_post_refuel_ticks', 0),
        'confidence_history': [
            {'timestamp': ts.isoformat() if hasattr(ts, 'isoformat') else str(ts), 'label': lbl, 'confidence': conf}
            for ts, lbl, conf in tracker.confidence_history[-50:]
        ],
        'active_event': tracker.event_record,
        'active_summary': active_summary,
        'has_history': len(state.event_history) > 0,
    }))

@app.route('/api/report')
@login_required
def api_report():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    with state.lock:
        if not state.processed_buffer:
            return "No data", 404
        df = pd.DataFrame(state.processed_buffer)
        df['Timestamp'] = df['Timestamp'].apply(lambda t: t.isoformat() if hasattr(t, 'isoformat') else str(t))
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='FuelSentinel_AI_Data', index=False)
    output.seek(0)
    return send_file(output, download_name=f'FuelSentinel_AI_Report_{car_id}.xlsx', as_attachment=True)

@app.route('/api/filter_status')
@login_required
def api_filter_status():
    car_id = request.args.get('car_id', session['user'].get('car_id') or current_car)
    error = authorize_car(car_id)
    if error:
        return error
    state = get_car_state(car_id)
    with state.lock:
        raw_recent = [p['fuel'] for p in list(state.data_buffer_raw)[-20:]]
        filtered_recent = [p['fuel'] for p in list(state.data_buffer_filtered)[-20:]]
        if not raw_recent:
            return jsonify({'error': 'No data'}), 404
        raw_std = float(np.std(raw_recent))
        filtered_std = float(np.std(filtered_recent))
        raw_mean = float(np.mean(raw_recent))
        filtered_mean = float(np.mean(filtered_recent))
        outlier_count = sum(1 for x in raw_recent if abs(x - raw_mean) / raw_std > 2.0) if raw_std > 0 else 0
        return jsonify(make_json_safe({
            'car_id': car_id,
            'buffer_size': len(filtered_recent),
            'current_spike_gate': state.noise_filter._spike_gate(),
            'current_window': state.noise_filter.current_window,
            'post_refuel_ticks': getattr(state.tracker, '_post_refuel_ticks', 0),
            'ema_value': state.delta_detector._ema,
            'raw': {'mean': raw_mean, 'std': raw_std, 'recent': raw_recent[-10:], 'outlier_count': outlier_count},
            'filtered': {
                'mean': filtered_mean, 'std': filtered_std, 'recent': filtered_recent[-10:],
                'noise_reduction': ((raw_std - filtered_std) / raw_std * 100) if raw_std > 0 else 0
            },
            'comparison': {
                'std_reduction': f"{(raw_std - filtered_std):.2f}L",
                'std_reduction_percent': f"{((raw_std - filtered_std) / raw_std * 100):.1f}%" if raw_std > 0 else "0%"
            }
        }))


if __name__ == '__main__':
    logger.info("🚀 FuelSentinel-AI V5 starting...")
    logger.info(f"   Model: {MODEL_PATH}")
    logger.info(f"   Device: {DEVICE}")
    logger.info(f"   Window ranges: {WINDOW_RANGES}")
    logger.info(f"   Thresholds: candidate={CANDIDATE_THRESH}, confirm_base={CONFIRM_THRESH}, finish={FINISH_THRESH}")
    print("=" * 60)
    print("FuelSentinel AI — V5")
    print("SUPABASE_URL:", SUPABASE_URL)
    print("SUPABASE_KEY:", SUPABASE_KEY[:20] + "..." if SUPABASE_KEY else "MISSING")
    print("=" * 60)
    fastapi_thread = threading.Thread(
        target=run_fastapi_backend,
        kwargs={"host": "127.0.0.1", "port": 8000},
        daemon=True,
    )
    fastapi_thread.start()
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)